"""
approvvigionamento_tab.py — Tab Approvvigionamento
Pre-cantiere: estrazione materiali da CME, stralci fornitori, confronto offerte.
Cantiere: tracciamento ordini, consegne, DDT.
"""

import io
import json
import pathlib
from datetime import date, timedelta

import streamlit as st


# ── Helper ─────────────────────────────────────────────────────────────────────

def _parse_importo(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace("€", "").strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


# ── Catalogo categorie materiali ───────────────────────────────────────────────

_CATEGORIE_MATERIALI = {
    "calcestruzzo":       {"icona": "🪨", "keywords": ["calcestruzzo", "cls", "cemento", "conglomerato cementizio", "getto"]},
    "acciaio":            {"icona": "⚙️", "keywords": ["acciaio", "ferro", "tondino", "armatura", "barra", "rete elettrosaldata", "HEA", "HEB", "IPE"]},
    "laterizi":           {"icona": "🧱", "keywords": ["laterizio", "mattone", "blocco", "forato", "pignatte", "tavelle"]},
    "legno":              {"icona": "🪵", "keywords": ["legno", "lamellare", "trave", "travetto", "perlinato", "tavolato"]},
    "impermeabilizzanti": {"icona": "💧", "keywords": ["impermeabilizzante", "guaina", "membrana", "bitume", "manto"]},
    "isolanti":           {"icona": "🌡️", "keywords": ["isolante", "polistirene", "lana di roccia", "cappotto", "EPS", "XPS"]},
    "intonaci":           {"icona": "🪣", "keywords": ["intonaco", "rasatura", "malta", "stucco", "rinzaffo"]},
    "pavimenti":          {"icona": "🔲", "keywords": ["pavimento", "piastrella", "ceramica", "gres", "parquet", "massetto"]},
    "infissi":            {"icona": "🚪", "keywords": ["infisso", "porta", "finestra", "serramento", "portone"]},
    "tubazioni":          {"icona": "🔧", "keywords": ["tubo", "tubazione", "raccordo", "pvc", "pe", "ghisa", "fognatura"]},
    "impianti_elettrici": {"icona": "⚡", "keywords": ["cavo", "conduttore", "quadro elettrico", "interruttore", "presa", "impianto elettrico"]},
    "inerti":             {"icona": "🪨", "keywords": ["sabbia", "ghiaia", "inerte", "pietrisco", "tout-venant"]},
    "verniciature":       {"icona": "🎨", "keywords": ["verniciatura", "pittura", "tinteggiatura", "vernice", "smalto"]},
    "altro":              {"icona": "📦", "keywords": []},
}


# ── Estrazione materiali da CME via Claude API ─────────────────────────────────

def _estrai_materiali_da_cme(testo_cme: str, api_key: str) -> list[dict]:
    """Analizza il CME ed estrae voci materiali con quantità, UM e prezzo."""
    import anthropic
    import uuid

    client = anthropic.Anthropic(api_key=api_key)
    prompt = f"""Sei un esperto di appalti pubblici italiani.
Analizza questo Computo Metrico Estimativo ed estrai TUTTE le voci
relative a materiali da acquistare (escludi le voci di sola manodopera).

Per ogni voce restituisci:
- categoria: una di (calcestruzzo|acciaio|laterizi|legno|impermeabilizzanti|isolanti|intonaci|pavimenti|infissi|tubazioni|impianti_elettrici|inerti|verniciature|altro)
- descrizione: descrizione completa della voce
- quantita: numero float
- um: unità di misura
- prezzo_unitario_gara: prezzo unitario dal CME (float)
- importo_gara: importo totale dal CME (float)

Restituisci SOLO array JSON valido, nessun testo aggiuntivo.

CME:
{testo_cme[:100000]}"""

    try:
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        voci = json.loads(raw.strip())
        for v in voci:
            v["id"] = f"cme_{uuid.uuid4().hex[:8]}"
            v.setdefault("offerte", [])
            v.setdefault("ordine", {
                "confermato": False,
                "data_conferma": None,
                "data_consegna_prevista": None,
                "data_consegna_effettiva": None,
                "quantita_consegnata": 0.0,
                "ddt_numeri": [],
                "note_consegna": "",
            })
        return voci
    except Exception:
        return []


# ── Generazione Excel stralcio per fornitore ───────────────────────────────────

def _genera_excel_stralcio(
    voci: list[dict],
    categoria_filtro: str,
    nome_progetto: str,
    mostra_prezzi: bool = False,
) -> bytes | None:
    """Genera Excel stralcio CME per richiesta offerta a fornitore."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter  # noqa: F401

        wb = Workbook()
        ws = wb.active
        ws.title = "Richiesta Offerta"

        # Intestazione
        ws.merge_cells("A1:F1")
        c_title = ws["A1"]
        c_title.value = f"RICHIESTA OFFERTA — {nome_progetto.upper()}"
        c_title.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c_title.fill = PatternFill("solid", fgColor="1E5099")
        c_title.alignment = Alignment(horizontal="center")

        ws["A2"].font = Font(bold=True)
        ws["A2"] = "Categoria:"
        ws["B2"] = categoria_filtro.upper() if categoria_filtro != "tutte" else "TUTTI I MATERIALI"

        ws["A3"].font = Font(bold=True)
        ws["A3"] = "Data richiesta:"
        ws["B3"] = date.today().strftime("%d/%m/%Y")

        headers = ["N°", "DESCRIZIONE", "Q.TÀ", "U.M.", "P.U. OFFERTA (€)", "IMPORTO (€)"]
        if not mostra_prezzi:
            headers[4] = "P.U. OFFERTA (€) — DA COMPILARE"
            headers[5] = "IMPORTO (€) — DA COMPILARE"

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E5099")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 18

        voci_filtrate = [
            v for v in voci
            if categoria_filtro == "tutte" or v.get("categoria") == categoria_filtro
        ]

        fill_alt = PatternFill("solid", fgColor="F8F8F8")
        fill_norm = PatternFill("solid", fgColor="FFFFFF")

        for row, (i, v) in enumerate(enumerate(voci_filtrate, 1), 6):
            fill = fill_alt if i % 2 == 0 else fill_norm
            ws.cell(row=row, column=1, value=i).fill = fill
            ws.cell(row=row, column=2, value=v.get("descrizione", "")).fill = fill
            ws.cell(row=row, column=3, value=v.get("quantita", 0)).fill = fill
            ws.cell(row=row, column=4, value=v.get("um", "")).fill = fill
            ws.cell(row=row, column=5, value=v.get("prezzo_unitario_gara", "") if mostra_prezzi else "").fill = fill
            ws.cell(row=row, column=6, value=v.get("importo_gara", "") if mostra_prezzi else "").fill = fill

        nota_row = 6 + len(voci_filtrate) + 1
        ws.cell(row=nota_row, column=1, value="NOTE:").font = Font(bold=True)
        ws.cell(row=nota_row + 1, column=1, value="Compilare le colonne P.U. e IMPORTO con i prezzi offerti e restituire firmato.")
        ws.cell(row=nota_row + 2, column=1, value=f"Offerta valida fino al: {(date.today() + timedelta(days=15)).strftime('%d/%m/%Y')}")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None


# ── Render principale ──────────────────────────────────────────────────────────

def render_approvvigionamento_tab(
    csa_data: dict,
    api_key: str,
    salva_fn,
    results_dir,
) -> None:
    nome_progetto = (
        f"{csa_data.get('tipo_lavori', 'Cantiere')} — "
        f"{csa_data.get('comune', '')} ({csa_data.get('provincia', '')})"
    )

    st.header("📦 Approvvigionamento Materiali")

    fase = st.radio(
        "Fase",
        ["🔍 Pre-cantiere", "🏗️ Cantiere"],
        horizontal=True,
        key="approv_fase_radio",
    )

    voci_cme: list[dict] = st.session_state.get("approv_voci_cme", [])

    # ════════════════════════════════════════════════════════════════════════════
    # FASE 1 — PRE-CANTIERE
    # ════════════════════════════════════════════════════════════════════════════
    if "Pre-cantiere" in fase:

        if voci_cme:
            tot_gara = sum(float(v.get("importo_gara", 0) or 0) for v in voci_cme)
            tot_offerte = sum(
                float(o.get("importo_totale", 0) or 0)
                for v in voci_cme
                for o in v.get("offerte", [])
                if o.get("selezionata")
            )
            margine = tot_gara - tot_offerte
            pct_margine = (margine / tot_gara * 100) if tot_gara > 0 else 0.0
            n_con_offerta = sum(
                1 for v in voci_cme
                if any(o.get("selezionata") for o in v.get("offerte", []))
            )

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("💰 Totale CME materiali", f"€ {tot_gara:,.2f}")
            col2.metric("📋 Offerte acquisite", f"€ {tot_offerte:,.2f}")
            col3.metric(
                "📈 Margine materiali",
                f"€ {margine:,.2f}",
                delta=f"{pct_margine:.1f}%",
                delta_color="normal" if margine >= 0 else "inverse",
            )
            col4.metric("Voci con offerta", f"{n_con_offerta}/{len(voci_cme)}")

        st.divider()

        # ── Step 1: Caricamento CME ───────────────────────────────────────────
        st.subheader("1️⃣ Caricamento CME Lavori")

        cme_bytes = st.session_state.get("piano_cme_bytes")
        cme_nome = st.session_state.get("piano_cme_nome", "")

        if cme_bytes:
            st.success(f"✅ CME: **{cme_nome}** (già caricato)")
            if st.button("🔄 Sostituisci CME", key="sost_cme_approv"):
                del st.session_state["piano_cme_bytes"]
                st.rerun()
        else:
            up = st.file_uploader(
                "Carica CME Lavori (PDF o Excel)",
                type=["pdf", "xlsx", "xls"],
                key="up_cme_approv",
            )
            if up and up.name != st.session_state.get("_cme_approv_nome_caricato"):
                st.session_state["piano_cme_bytes"] = up.read()
                st.session_state["piano_cme_nome"] = up.name
                st.session_state["_cme_approv_nome_caricato"] = up.name
                st.rerun()

        # ── Step 2: Estrazione materiali ──────────────────────────────────────
        st.subheader("2️⃣ Estrazione Materiali")

        if st.button(
            "🤖 Estrai materiali da CME",
            key="btn_estrai_mat",
            disabled=(not cme_bytes or not api_key),
        ):
            with st.spinner("Analisi CME in corso…"):
                from modules.csa_analyzer import _estrai_testo_pdf
                testo = _estrai_testo_pdf(cme_bytes)
                voci_cme = _estrai_materiali_da_cme(testo, api_key)
                st.session_state["approv_voci_cme"] = voci_cme
                salva_fn()
                st.success(f"✅ Estratte {len(voci_cme)} voci materiali")
                st.rerun()

        if not voci_cme:
            st.info("Nessun materiale estratto. Carica il CME e clicca 'Estrai materiali'.")
            return

        # ── Step 3: Genera stralci per fornitori ──────────────────────────────
        st.subheader("3️⃣ Genera Stralci per Fornitori")

        categorie_presenti = sorted(set(v.get("categoria", "altro") for v in voci_cme))
        col_s1, col_s2, col_s3 = st.columns(3)
        with col_s1:
            cat_sel = st.selectbox(
                "Categoria materiale",
                ["tutte"] + categorie_presenti,
                key="approv_cat_sel",
            )
        with col_s2:
            mostra_pz = st.checkbox(
                "Mostra prezzi di gara",
                value=False,
                key="approv_mostra_pz",
                help="⚠️ Normalmente NON mostrare i prezzi di gara al fornitore",
            )
        with col_s3:
            if st.button("📊 Genera Excel Stralcio", key="btn_gen_stralcio"):
                excel = _genera_excel_stralcio(voci_cme, cat_sel, nome_progetto, mostra_pz)
                if excel:
                    st.download_button(
                        "⬇️ Scarica Stralcio",
                        data=excel,
                        file_name=f"Stralcio_{cat_sel}_{date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_stralcio",
                    )
                else:
                    st.error("Errore generazione Excel — installa openpyxl")

        # ── Step 4: Inserimento offerte ───────────────────────────────────────
        st.subheader("4️⃣ Inserimento Offerte Fornitori")
        st.caption("Inserisci le offerte ricevute dai fornitori per ogni voce")

        cat_filtro = st.selectbox(
            "Filtra per categoria",
            ["tutte"] + categorie_presenti,
            key="approv_cat_filtro_off",
        )

        voci_filtrate = [
            v for v in voci_cme
            if cat_filtro == "tutte" or v.get("categoria") == cat_filtro
        ]

        for v in voci_filtrate:
            vid = v.get("id", "")
            offerte = v.get("offerte", [])
            offerta_sel = next((o for o in offerte if o.get("selezionata")), None)
            cat_info = _CATEGORIE_MATERIALI.get(v.get("categoria", "altro"), {"icona": "📦"})

            with st.container(border=True):
                col_a, col_b, col_c = st.columns([4, 2, 2])
                with col_a:
                    st.markdown(f"**{cat_info['icona']} {v.get('descrizione', '')}**")
                    st.caption(
                        f"Quantità: {v.get('quantita', 0)} {v.get('um', '')} "
                        f"— Prezzo gara: € {float(v.get('prezzo_unitario_gara', 0) or 0):,.2f}/{v.get('um', '')}"
                    )
                with col_b:
                    if offerta_sel:
                        risparmio = (
                            float(v.get("prezzo_unitario_gara", 0) or 0)
                            - float(offerta_sel.get("prezzo_unitario", 0) or 0)
                        ) * float(v.get("quantita", 0) or 0)
                        st.metric(
                            "Offerta selezionata",
                            f"€ {float(offerta_sel.get('importo_totale', 0) or 0):,.2f}",
                            delta=f"Risparmio € {risparmio:,.2f}",
                            delta_color="normal" if risparmio >= 0 else "inverse",
                        )
                    else:
                        st.warning("Nessuna offerta")
                with col_c:
                    if st.button("➕ Aggiungi offerta", key=f"add_off_{vid}"):
                        st.session_state[f"show_off_{vid}"] = True

                # Form nuova offerta
                if st.session_state.get(f"show_off_{vid}"):
                    with st.expander("Nuova offerta", expanded=True):
                        o_col1, o_col2 = st.columns(2)
                        with o_col1:
                            o_forn = st.text_input("Fornitore", key=f"off_forn_{vid}")
                            o_email = st.text_input("Email", key=f"off_email_{vid}")
                        with o_col2:
                            o_pu = st.number_input("Prezzo unitario (€)", min_value=0.0, key=f"off_pu_{vid}")
                            o_data = st.date_input("Data offerta", value=date.today(), key=f"off_data_{vid}", format="DD/MM/YYYY")
                        o_note = st.text_input("Note", key=f"off_note_{vid}")
                        o_imp = o_pu * float(v.get("quantita", 0) or 0)

                        if o_imp > 0:
                            risparmio_prev = float(v.get("importo_gara", 0) or 0) - o_imp
                            st.info(
                                f"Importo totale: **€ {o_imp:,.2f}** "
                                f"— Risparmio vs gara: **€ {risparmio_prev:,.2f}**"
                            )

                        if st.button("💾 Salva offerta", key=f"save_off_{vid}"):
                            for o in offerte:
                                o["selezionata"] = False
                            offerte.append({
                                "fornitore": o_forn,
                                "email": o_email,
                                "prezzo_unitario": o_pu,
                                "importo_totale": o_imp,
                                "data_offerta": str(o_data),
                                "note": o_note,
                                "selezionata": True,
                            })
                            v["offerte"] = offerte
                            st.session_state[f"show_off_{vid}"] = False
                            st.session_state["approv_voci_cme"] = voci_cme
                            salva_fn()
                            st.rerun()

    # ════════════════════════════════════════════════════════════════════════════
    # FASE 2 — CANTIERE
    # ════════════════════════════════════════════════════════════════════════════
    else:
        st.subheader("🏗️ Gestione Ordini e Consegne")

        if not voci_cme:
            st.info("Nessun materiale estratto. Completa prima la fase Pre-cantiere.")
            return

        tot_voci = len(voci_cme)
        conf_ordini = sum(1 for v in voci_cme if v.get("ordine", {}).get("confermato"))
        consegnati = sum(
            1 for v in voci_cme
            if float(v.get("ordine", {}).get("quantita_consegnata", 0) or 0)
            >= float(v.get("quantita", 0) or 0) > 0
        )
        parziali = sum(
            1 for v in voci_cme
            if 0 < float(v.get("ordine", {}).get("quantita_consegnata", 0) or 0)
            < float(v.get("quantita", 0) or 0)
        )

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("📦 Totale voci", tot_voci)
        col2.metric("✅ Ordini confermati", conf_ordini)
        col3.metric("🚚 Consegnati", consegnati)
        col4.metric("⏳ Consegne parziali", parziali)

        st.divider()

        filtro_stato = st.selectbox(
            "Filtra per stato",
            ["tutti", "da ordinare", "ordinato", "consegna parziale", "consegnato"],
            key="approv_filtro_stato",
        )

        for v in voci_cme:
            ordine = v.get("ordine", {})
            qta_totale = float(v.get("quantita", 0) or 0)
            qta_consegnata = float(ordine.get("quantita_consegnata", 0) or 0)
            confermato = ordine.get("confermato", False)

            if qta_totale > 0 and qta_consegnata >= qta_totale:
                stato = "consegnato"
                stato_emoji = "✅"
            elif qta_consegnata > 0:
                stato = "consegna parziale"
                stato_emoji = "⏳"
            elif confermato:
                stato = "ordinato"
                stato_emoji = "📋"
            else:
                stato = "da ordinare"
                stato_emoji = "🔴"

            if filtro_stato != "tutti" and stato != filtro_stato:
                continue

            vid = v.get("id", "")
            cat_info = _CATEGORIE_MATERIALI.get(v.get("categoria", "altro"), {"icona": "📦"})

            with st.container(border=True):
                col_a, col_b, col_c = st.columns([4, 3, 1])
                with col_a:
                    st.markdown(f"**{cat_info['icona']} {v.get('descrizione', '')}**")
                    offerta_sel = next(
                        (o for o in v.get("offerte", []) if o.get("selezionata")), None
                    )
                    if offerta_sel:
                        st.caption(
                            f"Fornitore: {offerta_sel.get('fornitore', '—')} "
                            f"— € {float(offerta_sel.get('importo_totale', 0) or 0):,.2f}"
                        )
                with col_b:
                    progress_val = min(qta_consegnata / qta_totale, 1.0) if qta_totale > 0 else 0.0
                    st.progress(
                        progress_val,
                        text=f"Consegnato: {qta_consegnata}/{qta_totale} {v.get('um', '')}",
                    )
                with col_c:
                    st.markdown(f"## {stato_emoji}")

                if st.button("📝 Aggiorna consegna", key=f"upd_cons_{vid}"):
                    st.session_state[f"show_cons_{vid}"] = True

                if st.session_state.get(f"show_cons_{vid}"):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        nuova_qta = st.number_input(
                            "Quantità consegnata",
                            min_value=0.0,
                            value=qta_consegnata,
                            key=f"cons_qta_{vid}",
                        )
                        data_cons = st.date_input(
                            "Data consegna",
                            value=date.today(),
                            key=f"cons_data_{vid}",
                            format="DD/MM/YYYY",
                        )
                    with c2:
                        ddt = st.text_input(
                            "N° DDT",
                            key=f"cons_ddt_{vid}",
                            placeholder="DDT-001/2026",
                        )
                        note_cons = st.text_input("Note consegna", key=f"cons_note_{vid}")
                    with c3:
                        confermato_new = st.checkbox(
                            "Ordine confermato",
                            value=confermato,
                            key=f"cons_conf_{vid}",
                        )
                        data_prev = st.date_input(
                            "Consegna prevista",
                            value=date.today(),
                            key=f"cons_prev_{vid}",
                            format="DD/MM/YYYY",
                        )

                    if st.button("💾 Salva", key=f"save_cons_{vid}"):
                        ddt_list = list(ordine.get("ddt_numeri", []))
                        if ddt:
                            ddt_list.append(ddt)
                        v["ordine"] = {
                            "confermato": confermato_new,
                            "data_conferma": str(date.today()) if confermato_new else None,
                            "data_consegna_prevista": str(data_prev),
                            "data_consegna_effettiva": str(data_cons) if nuova_qta > 0 else None,
                            "quantita_consegnata": nuova_qta,
                            "ddt_numeri": ddt_list,
                            "note_consegna": note_cons,
                        }
                        st.session_state[f"show_cons_{vid}"] = False
                        st.session_state["approv_voci_cme"] = voci_cme
                        salva_fn()
                        st.rerun()
