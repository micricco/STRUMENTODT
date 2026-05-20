"""
pianificazione_tab.py — Tab Pianificazione per app DTC.

Layout 2 colonne: Cronoprogramma (sx) + Documenti CME (dx).
KPI in cima, associazione fasi→voci, generazione Excel
tramite pianificazione_excel.genera_excel_pianificazione().
"""

import io
import json
import re
import pathlib

import pandas as pd
import streamlit as st
import anthropic

from modules.log_manager import aggiungi_log, aggiungi_al_diario
from modules.csa_analyzer import _estrai_testo_pdf
from pianificazione_excel import genera_excel_pianificazione

_MODELLO = "claude-haiku-4-5-20251001"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_json_risposta(raw: str) -> list | dict:
    raw = raw.strip()
    if raw.startswith("```"):
        lines = raw.split("\n")
        inner = lines[1:-1] if lines and lines[-1].strip() == "```" else lines[1:]
        raw = "\n".join(inner)
    try:
        return json.loads(raw)
    except Exception:
        m = re.search(r"(\[[\s\S]*\]|\{[\s\S]*\})", raw)
        if m:
            try:
                return json.loads(m.group(1))
            except Exception:
                pass
    return []


def _init_piano() -> dict:
    return {
        "fasi": [],
        "voci_cme": [],
        "voci_sicurezza": [],
        "totale_cme": 0.0,
        "totale_sicurezza": 0.0,
        "associazioni": {},
        "_gantt_file_id": None,
        "_cme_file_ids": [],
        "_sicurezza_file_id": None,
        "_associazione_auto_fatta": False,
    }


def _conta_pagine_pdf(fb: bytes) -> int:
    try:
        import fitz
        doc = fitz.open(stream=fb, filetype="pdf")
        n = doc.page_count
        doc.close()
        return n
    except Exception:
        return 0


def _get_pagine_cached(f) -> int:
    """Conta pagine di un UploadedFile, con cache per evitare riletture."""
    cache = st.session_state.setdefault("pian_pg_cache", {})
    fid = f"{f.name}_{f.size}"
    if fid not in cache:
        fb = f.read()
        f.seek(0)
        cache[fid] = _conta_pagine_pdf(fb)
    return cache[fid]


# ── Estrazione Gantt ──────────────────────────────────────────────────────────

def _estrai_fasi_gantt_excel(file_bytes: bytes) -> list[dict]:
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        st.error(f"Errore lettura Excel: {e}")
        return []

    col_map: dict[str, str] = {}
    for col in df.columns:
        cl = str(col).lower().strip()
        if any(k in cl for k in ("attivit", "nome", "fase", "lavoraz", "descr")):
            col_map.setdefault("nome", col)
        elif any(k in cl for k in ("durata", "duration")):
            col_map.setdefault("durata", col)
        elif any(k in cl for k in ("inizio", "start")):
            col_map.setdefault("inizio", col)
        elif any(k in cl for k in ("fine", "end", "finish")):
            col_map.setdefault("fine", col)

    nome_col = col_map.get("nome", df.columns[0] if len(df.columns) > 0 else None)
    if nome_col is None:
        return []

    fasi = []
    for _, row in df.iterrows():
        nome = str(row.get(nome_col, "") or "").strip()
        if not nome or nome.lower() in ("nan", "none", ""):
            continue

        def _int_safe(col_key):
            col = col_map.get(col_key)
            if not col:
                return None
            try:
                v = row.get(col)
                return int(v) if v is not None and str(v) not in ("nan", "None", "") else None
            except Exception:
                return None

        inizio = _int_safe("inizio")
        fine = _int_safe("fine")
        durata_gg = _int_safe("durata")

        inizio = inizio or 1
        fine = fine or inizio
        if not durata_gg:
            durata_gg = max(1, (fine - inizio + 1) * 5)

        fasi.append({
            "nome": nome,
            "durata_giorni": durata_gg,
            "settimana_inizio": inizio,
            "settimana_fine": fine,
        })
    return fasi


def _estrai_fasi_gantt_pdf(testo: str, api_key: str) -> list[dict]:
    client = anthropic.Anthropic(api_key=api_key)
    prompt = (
        "Estrai le fasi del cronoprogramma in JSON.\n"
        "Rispondi SOLO con JSON valido.\n"
        '[{"nome": "ORGANIZZAZIONE CANTIERE",'
        '"durata_giorni": 5,'
        '"settimana_inizio": 1,'
        '"settimana_fine": 1}]\n\n'
        f"TESTO:\n{testo[:50_000]}"
    )
    try:
        resp = client.messages.create(
            model=_MODELLO, max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        result = _parse_json_risposta(resp.content[0].text)
        return result if isinstance(result, list) else []
    except Exception as e:
        st.warning(f"Errore estrazione Gantt da PDF: {e}")
        return []


# ── Estrazione CME e Sicurezza ────────────────────────────────────────────────

def _estrai_voci_cme(testo: str, api_key: str) -> list[dict]:
    client = anthropic.Anthropic(api_key=api_key)
    prompt = (
        "Estrai le voci del computo metrico in JSON.\n"
        "Rispondi SOLO con JSON valido, nessun testo.\n"
        '[{"numero": "1/1",'
        '"categoria": "DEMOLIZIONI",'
        '"descrizione": "Rimozione manto copertura...",'
        '"quantita": 180.00,'
        '"unita_misura": "mq",'
        '"prezzo_unitario": 19.66,'
        '"importo_totale": 3538.80,'
        '"tipo": "cme"}]\n\n'
        f"TESTO:\n{testo[:60_000]}"
    )
    try:
        resp = client.messages.create(
            model=_MODELLO, max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        result = _parse_json_risposta(resp.content[0].text)
        return result if isinstance(result, list) else []
    except Exception as e:
        st.warning(f"Errore estrazione CME: {e}")
        return []


def _estrai_voci_sicurezza(testo: str, api_key: str) -> list[dict]:
    client = anthropic.Anthropic(api_key=api_key)
    prompt = (
        "Estrai le voci degli oneri per la sicurezza in JSON.\n"
        "Ogni voce ha nota 'NON SOGGETTO A RIBASSO'.\n"
        "Rispondi SOLO con JSON valido, nessun testo.\n"
        '[{"numero": "1",'
        '"categoria": "SICUREZZA",'
        '"descrizione": "Ponteggio telaio prefabbricato h>10m...",'
        '"quantita": 648.0,'
        '"unita_misura": "mq",'
        '"prezzo_unitario": 21.53,'
        '"importo_totale": 13951.44,'
        '"tipo": "sicurezza",'
        '"nota": "NON SOGGETTO A RIBASSO"}]\n\n'
        f"TESTO:\n{testo[:60_000]}"
    )
    try:
        resp = client.messages.create(
            model=_MODELLO, max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        result = _parse_json_risposta(resp.content[0].text)
        return result if isinstance(result, list) else []
    except Exception as e:
        st.warning(f"Errore estrazione Oneri Sicurezza: {e}")
        return []


# ── Associazione automatica ───────────────────────────────────────────────────

def _associa_fasi_voci(fasi: list[dict], voci: list[dict], api_key: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    json_fasi = json.dumps(
        [{"nome": f.get("nome", ""), "durata_giorni": f.get("durata_giorni", 0)}
         for f in fasi],
        ensure_ascii=False,
    )
    json_voci = json.dumps(
        [{"id": str(i), "numero": v.get("numero", ""), "descrizione": v.get("descrizione", "")}
         for i, v in enumerate(voci[:150])],
        ensure_ascii=False,
    )
    prompt = (
        f"Hai queste fasi del cronoprogramma:\n{json_fasi}\n\n"
        f"E queste voci di computo metrico:\n{json_voci}\n\n"
        "Associa ogni voce alla fase più probabile basandoti sulla descrizione della lavorazione.\n"
        "Rispondi SOLO con JSON:\n"
        '{"FASE 1 – ORGANIZZAZIONE": ["0","1","2"],'
        '"FASE 2 – DEMOLIZIONI": ["3","4","5"]}'
    )
    try:
        resp = client.messages.create(
            model=_MODELLO, max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        result = _parse_json_risposta(resp.content[0].text)
        return result if isinstance(result, dict) else {}
    except Exception as e:
        st.warning(f"Errore associazione automatica: {e}")
        return {}


# ── Generazione Excel Cruscotto per Fase ─────────────────────────────────────

def _genera_excel_cruscotto(pianificazione_data: dict, csa_data: dict) -> bytes | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        fasi = pianificazione_data.get("fasi", [])
        voci_cme_all = pianificazione_data.get("voci_cme", [])
        voci_sic_all = pianificazione_data.get("voci_sicurezza", [])
        associazioni = pianificazione_data.get("associazioni", {})
        n_cme = len(voci_cme_all)

        # voce_index → nome_fase
        voce_to_fase_nome: dict[int, str] = {}
        for nome_fase_k, ids in associazioni.items():
            for vid in ids:
                try:
                    voce_to_fase_nome[int(vid)] = nome_fase_k
                except (ValueError, TypeError):
                    pass

        nome_to_fi = {
            str(f.get("nome", f"Fase {fi+1}")): fi
            for fi, f in enumerate(fasi)
        }

        fase_cme: dict[int, list] = {fi: [] for fi in range(len(fasi))}
        fase_sic: dict[int, list] = {fi: [] for fi in range(len(fasi))}

        for i, v in enumerate(voci_cme_all):
            fi = nome_to_fi.get(voce_to_fase_nome.get(i, ""))
            if fi is not None:
                fase_cme[fi].append(v)

        for i, v in enumerate(voci_sic_all):
            fi = nome_to_fi.get(voce_to_fase_nome.get(n_cme + i, ""))
            if fi is not None:
                fase_sic[fi].append(v)

        # ── Workbook ──────────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Cruscotto per Fase"

        # Stili
        fill_blu_scuro = PatternFill("solid", fgColor="00467F")
        fill_blu_medio = PatternFill("solid", fgColor="1E5099")
        fill_rosso     = PatternFill("solid", fgColor="FFC8C8")
        fill_grigio_ch = PatternFill("solid", fgColor="DCDCDC")
        fill_grigio_lt = PatternFill("solid", fgColor="F0F0F0")
        fill_giallo    = PatternFill("solid", fgColor="FFFFB4")
        fill_oro       = PatternFill("solid", fgColor="FFD700")
        fill_bianco    = PatternFill("solid", fgColor="FFFFFF")
        fill_alt       = PatternFill("solid", fgColor="F8F8F8")

        font_hdr  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_fase = Font(name="Arial", size=9,  bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=9,  bold=True)
        font_it   = Font(name="Arial", size=9,  italic=True)
        font_norm = Font(name="Arial", size=9)
        font_tot  = Font(name="Arial", size=9,  bold=True)
        font_gen  = Font(name="Arial", size=10, bold=True)

        thin   = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        al_c = Alignment(horizontal="center", vertical="center")
        al_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        al_r = Alignment(horizontal="right",  vertical="center")

        fmt_euro = '#,##0.00'

        # Larghezze colonne: A=5 B=55 C=8 D=8 E=7 F=7 G=7 H=12 I=14 J=14
        for ci, w in enumerate([5, 55, 8, 8, 7, 7, 7, 12, 14, 14], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        def _apply(cell, fill=None, font=None, align=None, fmt=None):
            if fill:  cell.fill = fill
            if font:  cell.font = font
            if align: cell.alignment = align
            if fmt:   cell.number_format = fmt
            cell.border = border

        # ── Header colonne (riga 1) ───────────────────────────────────────────
        headers = [
            "N°", "DESIGNAZIONE DEI LAVORI", "Q.TÀ", "U.M.",
            "S.INI", "S.FIN", "GG", "P.U. (€)", "IMP. CME (€)", "IMP. SIC. (€)",
        ]
        ws.row_dimensions[1].height = 22
        for ci, h in enumerate(headers, 1):
            _apply(ws.cell(row=1, column=ci, value=h),
                   fill=fill_blu_medio, font=font_hdr, align=al_c)
        ws.freeze_panes = "A2"

        row = 2
        fase_tot_rows_cme: list[int] = []  # riga totale fase con valore col I
        fase_tot_rows_sic: list[int] = []  # riga totale fase con valore col J

        for fi, fase in enumerate(fasi):
            nome_fase = str(fase.get("nome", f"Fase {fi+1}"))
            s_ini = fase.get("settimana_inizio", "")
            s_fin = fase.get("settimana_fine", "")
            gg    = fase.get("durata_giorni", "")
            cme_voci = fase_cme.get(fi, [])
            sic_voci = fase_sic.get(fi, [])

            # ── Header fase ──────────────────────────────────────────────────
            ws.row_dimensions[row].height = 18
            ws.merge_cells(f"A{row}:H{row}")
            _apply(ws.cell(row=row, column=1,
                           value=f"FASE {fi+1} – {nome_fase.upper()}"),
                   fill=fill_blu_scuro, font=font_fase, align=al_l)
            _apply(ws.cell(row=row, column=9,
                           value=f"Sett.{s_ini}÷{s_fin}"),
                   fill=fill_blu_scuro, font=font_fase, align=al_c)
            _apply(ws.cell(row=row, column=10,
                           value=f"{gg} gg"),
                   fill=fill_blu_scuro, font=font_fase, align=al_c)
            row += 1

            # ── Voci sicurezza ────────────────────────────────────────────────
            sic_range_start = row
            if sic_voci:
                ws.merge_cells(f"A{row}:J{row}")
                _apply(ws.cell(row=row, column=1,
                               value="   🔴  ONERI PER LA SICUREZZA – NON SOGGETTI A RIBASSO"),
                       fill=fill_rosso, font=font_bold, align=al_l)
                for c in range(2, 11):
                    ws.cell(row=row, column=c).fill = fill_rosso
                    ws.cell(row=row, column=c).border = border
                row += 1

                for j, v in enumerate(sic_voci):
                    fill = fill_bianco if j % 2 == 0 else fill_alt
                    ws.row_dimensions[row].height = 28
                    vals = [
                        j + 1,
                        v.get("descrizione", ""),
                        v.get("quantita", ""),
                        v.get("unita_misura", ""),
                        s_ini, s_fin, gg,
                        float(v.get("prezzo_unitario") or 0),
                        None,
                        float(v.get("importo_totale") or 0),
                    ]
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        _apply(cell, fill=fill, font=font_norm,
                               align=(al_c if ci in (1,3,4,5,6,7) else
                                      al_r if ci in (8,9,10) else al_l),
                               fmt=(fmt_euro if ci in (8,10) and val else None))
                    row += 1
            sic_range_end = row - 1

            # ── Voci CME raggruppate per categoria ───────────────────────────
            cme_range_start = row
            if cme_voci:
                ws.merge_cells(f"A{row}:J{row}")
                _apply(ws.cell(row=row, column=1,
                               value="   ▸  LAVORI A MISURA – COMPUTO METRICO ESTIMATIVO"),
                       fill=fill_grigio_ch, font=font_bold, align=al_l)
                for c in range(2, 11):
                    ws.cell(row=row, column=c).fill = fill_grigio_ch
                    ws.cell(row=row, column=c).border = border
                row += 1

                categorie: dict[str, list] = {}
                for v in cme_voci:
                    cat = str(v.get("categoria") or "LAVORAZIONI").upper()
                    categorie.setdefault(cat, []).append(v)

                n_prog = 1
                for cat_name, cat_voci in categorie.items():
                    ws.merge_cells(f"A{row}:J{row}")
                    _apply(ws.cell(row=row, column=1,
                                   value=f"      › {cat_name}"),
                           fill=fill_grigio_lt, font=font_it, align=al_l)
                    for c in range(2, 11):
                        ws.cell(row=row, column=c).fill = fill_grigio_lt
                        ws.cell(row=row, column=c).border = border
                    row += 1

                    for j, v in enumerate(cat_voci):
                        fill = fill_bianco if j % 2 == 0 else fill_alt
                        ws.row_dimensions[row].height = 28
                        vals = [
                            n_prog,
                            v.get("descrizione", ""),
                            v.get("quantita", ""),
                            v.get("unita_misura", ""),
                            s_ini, s_fin, gg,
                            float(v.get("prezzo_unitario") or 0),
                            float(v.get("importo_totale") or 0),
                            None,
                        ]
                        for ci, val in enumerate(vals, 1):
                            cell = ws.cell(row=row, column=ci, value=val)
                            _apply(cell, fill=fill, font=font_norm,
                                   align=(al_c if ci in (1,3,4,5,6,7) else
                                          al_r if ci in (8,9,10) else al_l),
                                   fmt=(fmt_euro if ci in (8,9) and val else None))
                        row += 1
                        n_prog += 1
            cme_range_end = row - 1

            # ── Riga totale fase ──────────────────────────────────────────────
            ws.merge_cells(f"A{row}:H{row}")
            _apply(ws.cell(row=row, column=1,
                           value=f"TOTALE FASE {fi+1} – {nome_fase.upper()}"),
                   fill=fill_giallo, font=font_tot, align=al_l)

            cme_formula = (
                f"=SUM(I{cme_range_start}:I{cme_range_end})"
                if cme_voci else 0
            )
            sic_formula = (
                f"=SUM(J{sic_range_start}:J{sic_range_end})"
                if sic_voci else 0
            )

            cell_ci = ws.cell(row=row, column=9, value=cme_formula)
            _apply(cell_ci, fill=fill_giallo, font=font_tot, align=al_r,
                   fmt=fmt_euro)
            cell_cj = ws.cell(row=row, column=10, value=sic_formula)
            _apply(cell_cj, fill=fill_giallo, font=font_tot, align=al_r,
                   fmt=fmt_euro)

            if cme_voci:
                fase_tot_rows_cme.append(row)
            if sic_voci:
                fase_tot_rows_sic.append(row)
            row += 1

        # ── Totali generali ───────────────────────────────────────────────────
        row += 1

        def _riga_totale(r, label, col_i_val, col_j_val, fill, font):
            ws.merge_cells(f"A{r}:H{r}")
            _apply(ws.cell(row=r, column=1, value=label),
                   fill=fill, font=font, align=al_l)
            ci_cell = ws.cell(row=r, column=9, value=col_i_val)
            _apply(ci_cell, fill=fill, font=font, align=al_r, fmt=fmt_euro)
            cj_cell = ws.cell(row=r, column=10, value=col_j_val)
            _apply(cj_cell, fill=fill, font=font, align=al_r, fmt=fmt_euro)
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = fill
                ws.cell(row=r, column=c).border = border

        tot_lavori_formula = (
            "=SUM(" + ",".join(f"I{r}" for r in fase_tot_rows_cme) + ")"
            if fase_tot_rows_cme else 0
        )
        tot_sic_formula = (
            "=SUM(" + ",".join(f"J{r}" for r in fase_tot_rows_sic) + ")"
            if fase_tot_rows_sic else 0
        )

        _riga_totale(row, "TOTALE GENERALE LAVORI",
                     tot_lavori_formula, "", fill_giallo, font_tot)
        tot_lavori_row = row
        row += 1

        _riga_totale(row, "TOTALE GENERALE SICUREZZA",
                     "", tot_sic_formula, fill_giallo, font_tot)
        tot_sic_row = row
        row += 1

        totale_complessivo = f"=I{tot_lavori_row}+J{tot_sic_row}"
        _riga_totale(row, "TOTALE COMPLESSIVO",
                     totale_complessivo, "", fill_oro, font_gen)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    except Exception as e:
        st.error(f"Errore generazione Excel Cruscotto: {e}")
        return None


# ── Rilevamento documenti già caricati nel tab Documenti ─────────────────────

_KEYWORDS_PIAN: dict[str, list[str]] = {
    "gantt":     ["cronoprogramma", "programma lavori", "gantt", "programma dei lavori"],
    "cme":       ["computo metrico", "cme", "computo estimativo", "computo metrico estimativo"],
    "elenco":    ["elenco prezzi", "elenco dei prezzi", "listino prezzi"],
    "sicurezza": ["oneri sicurezza", "oneri per la sicurezza", "psc",
                  "piano di sicurezza", "misure sicurezza"],
}

_LABEL_PIAN: dict[str, str] = {
    "gantt":     "📅 Cronoprogramma",
    "cme":       "📋 Computo Metrico Estimativo",
    "elenco":    "💰 Elenco Prezzi",
    "sicurezza": "🦺 Oneri per la Sicurezza",
}


def _trova_doc_da_elaborati() -> dict[str, dict]:
    """Scansiona doc_elaborati cercando documenti utili alla pianificazione."""
    doc_elaborati = st.session_state.get("doc_elaborati", {})
    trovati: dict[str, dict] = {}
    for cat_docs in doc_elaborati.values():
        for doc in (cat_docs or []):
            if not doc.get("path"):
                continue
            testo = (doc.get("titolo", "") + " " + doc.get("codice", "")).lower()
            for tipo, kws in _KEYWORDS_PIAN.items():
                if tipo not in trovati and any(kw in testo for kw in kws):
                    trovati[tipo] = doc
                    break
    return trovati


# ── Render principale ─────────────────────────────────────────────────────────

def render_pianificazione_tab(
    csa_data: dict,
    api_key: str,
    salva_fn,
    results_dir: pathlib.Path,
) -> None:

    if "pianificazione" not in st.session_state:
        st.session_state.pianificazione = _init_piano()
    piano = st.session_state.pianificazione

    nome_progetto = (
        f"{csa_data.get('tipo_lavori', 'Cantiere')} — "
        f"{csa_data.get('comune', '')} ({csa_data.get('provincia', '')})"
    )

    # ── Documenti flaggati dal tab Documenti ──────────────────────────────────
    docs_flaggati = st.session_state.get("docs_per_pianificazione", {})

    _RUOLI_MAP = {
        "📅 Cronoprogramma": "gantt",
        "📋 CME":            "cme",
        "💰 Elenco Prezzi":  "elenco",
        "🦺 Oneri Sicurezza": "sicurezza",
    }

    if docs_flaggati:
        st.info(
            f"📂 **{len(docs_flaggati)} documento/i** contrassegnati per la pianificazione "
            "— seleziona il ruolo e clicca ✅ Usa:"
        )
        for codice, doc in docs_flaggati.items():
            c1, c2, c3 = st.columns([3, 2, 1])
            with c1:
                st.caption(f"📄 **{doc.get('titolo') or doc.get('nome', '')}**")
            with c2:
                ruolo = st.selectbox(
                    "Ruolo",
                    ["— seleziona —"] + list(_RUOLI_MAP.keys()),
                    key=f"assegna_{codice}",
                    label_visibility="collapsed",
                )
            with c3:
                if st.button("✅ Usa", key=f"usa_{codice}"):
                    tipo = _RUOLI_MAP.get(ruolo)
                    if tipo and doc.get("path"):
                        try:
                            fb = pathlib.Path(doc["path"]).read_bytes()
                            st.session_state[f"piano_{tipo}_bytes"] = fb
                            st.session_state[f"piano_{tipo}_nome"] = doc.get("nome", "")
                            st.session_state[f"piano_{tipo}_path"] = doc.get("path", "")
                            piano["_associazione_auto_fatta"] = False
                            st.rerun()
                        except Exception as e:
                            st.error(f"Errore lettura file: {e}")
                    elif not tipo:
                        st.warning("Seleziona un ruolo prima.")
                    else:
                        st.warning("File non disponibile.")
        st.divider()
    else:
        st.caption(
            "💡 Vai nel tab **Documenti** e attiva il toggle **📋 Pianificazione** "
            "sui file da usare qui."
        )

    # ── KPI sempre visibili in cima ───────────────────────────────────────────
    voci_cme = piano.get("voci_cme", [])
    voci_sicurezza = piano.get("voci_sicurezza", [])
    fasi = piano.get("fasi", [])

    tot_cme = sum(float(v.get("importo_totale") or 0) for v in voci_cme)
    tot_sic = sum(float(v.get("importo_totale") or 0) for v in voci_sicurezza)
    tot_appalto = tot_cme + tot_sic
    durata_tot = sum(int(f.get("durata_giorni") or 0) for f in fasi)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 Totale CME", f"€ {tot_cme:,.2f}")
    col2.metric("🔴 Oneri Sicurezza", f"€ {tot_sic:,.2f}")
    col3.metric("📊 Totale Appalto", f"€ {tot_appalto:,.2f}")
    col4.metric("📅 Durata totale", f"{durata_tot} gg" if durata_tot else "—")

    st.divider()

    # ── 2 colonne affiancate ──────────────────────────────────────────────────
    col_sx, col_dx = st.columns([1, 1])

    # ════════════════════════════════════════════════════════════════════════
    # COLONNA SINISTRA — Cronoprogramma
    # ════════════════════════════════════════════════════════════════════════
    with col_sx:
        st.subheader("📅 Cronoprogramma")

        gantt_b_imp = st.session_state.get("piano_gantt_bytes")
        gantt_n_imp = st.session_state.get("piano_gantt_nome", "")

        def _processa_gantt(fb: bytes, nome: str, fid: str) -> None:
            fasi_e: list[dict] = []
            if nome.lower().endswith((".xlsx", ".xls")):
                fasi_e = _estrai_fasi_gantt_excel(fb)
            else:
                if not api_key:
                    st.error("API key mancante per analizzare PDF Gantt.")
                    return
                with st.spinner("Estrazione fasi da PDF…"):
                    fasi_e = _estrai_fasi_gantt_pdf(_estrai_testo_pdf(fb), api_key)
            if fasi_e:
                piano["fasi"] = fasi_e
                piano["_gantt_file_id"] = fid
                piano["_associazione_auto_fatta"] = False
                aggiungi_log("Gantt caricato", f"{nome} — {len(fasi_e)} fasi", "Pianificazione")
                aggiungi_al_diario(
                    f"Cronoprogramma caricato: {nome} — {len(fasi_e)} fasi",
                    "🟡 Adempimento Organizzativo",
                )
                salva_fn()
                st.rerun()
            elif not piano.get("fasi"):
                st.warning("Nessuna fase estratta. Usa l'inserimento manuale.")
            else:
                piano["_gantt_file_id"] = fid

        if gantt_b_imp:
            st.success(f"✅ Cronoprogramma: **{gantt_n_imp}** (importato da Documenti)")
            _, col_sost_g = st.columns([3, 1])
            with col_sost_g:
                if st.button("🔄 Sostituisci", key="sostituisci_gantt"):
                    for k in ("piano_gantt_bytes", "piano_gantt_nome", "piano_gantt_path"):
                        st.session_state.pop(k, None)
                    st.rerun()
            fid_imp_g = f"piano_gantt_{gantt_n_imp}"
            if fid_imp_g != piano.get("_gantt_file_id"):
                _processa_gantt(gantt_b_imp, gantt_n_imp, fid_imp_g)
        else:
            file_gantt = st.file_uploader(
                "Carica Cronoprogramma",
                type=["pdf", "xlsx", "xls"],
                key="upload_gantt",
            )
            if file_gantt:
                fid = f"{file_gantt.name}_{file_gantt.size}"
                if fid != piano.get("_gantt_file_id"):
                    _processa_gantt(file_gantt.read(), file_gantt.name, fid)

        if piano.get("fasi"):
            df_fasi = pd.DataFrame(piano["fasi"])
            for c in ["nome", "durata_giorni", "settimana_inizio", "settimana_fine"]:
                if c not in df_fasi.columns:
                    df_fasi[c] = "" if c == "nome" else 1
            edited = st.data_editor(
                df_fasi[["nome", "durata_giorni", "settimana_inizio", "settimana_fine"]],
                use_container_width=True,
                num_rows="dynamic",
                key="pian_fasi_de",
                column_config={
                    "nome": st.column_config.TextColumn("Nome Fase", width="large"),
                    "durata_giorni": st.column_config.NumberColumn("Durata (gg)", min_value=1),
                    "settimana_inizio": st.column_config.NumberColumn("Sett. Inizio", min_value=1),
                    "settimana_fine": st.column_config.NumberColumn("Sett. Fine", min_value=1),
                },
            )
            if edited is not None:
                aggiornate = edited.fillna("").to_dict("records")
                if aggiornate != piano["fasi"]:
                    piano["fasi"] = aggiornate
                    piano["_associazione_auto_fatta"] = False
                    salva_fn()
        else:
            st.info("Nessun cronoprogramma caricato.")
            if st.button("➕ Inserisci fasi manualmente", key="pian_gantt_manual"):
                piano["fasi"] = [
                    {"nome": "FASE 1 – ORGANIZZAZIONE CANTIERE",
                     "durata_giorni": 5, "settimana_inizio": 1, "settimana_fine": 1},
                ]
                piano["_gantt_file_id"] = "manual"
                salva_fn()
                st.rerun()

    # ════════════════════════════════════════════════════════════════════════
    # COLONNA DESTRA — Documenti di Computo
    # ════════════════════════════════════════════════════════════════════════
    with col_dx:
        st.subheader("📋 Documenti di Computo")

        # ── CME ──────────────────────────────────────────────────────────────
        cme_b_imp = st.session_state.get("piano_cme_bytes")
        cme_n_imp = st.session_state.get("piano_cme_nome", "")
        if cme_b_imp:
            st.success(f"✅ CME: **{cme_n_imp}** (importato da Documenti)")
            if st.button("🔄 Sostituisci CME", key="sostituisci_cme"):
                for k in ("piano_cme_bytes", "piano_cme_nome", "piano_cme_path"):
                    st.session_state.pop(k, None)
                st.rerun()
            files_cme = []
        else:
            files_cme = st.file_uploader(
                "Computo Metrico Estimativo (uno o più PDF)",
                type=["pdf"],
                accept_multiple_files=True,
                key="upload_cme",
            )
            for f in (files_cme or []):
                st.success(f"✅ {f.name} — {_get_pagine_cached(f)} pagine")

        # ── Elenco prezzi ─────────────────────────────────────────────────────
        elenco_b_imp = st.session_state.get("piano_elenco_bytes")
        elenco_n_imp = st.session_state.get("piano_elenco_nome", "")
        if elenco_b_imp:
            st.success(f"✅ Elenco Prezzi: **{elenco_n_imp}** (importato da Documenti)")
            if st.button("🔄 Sostituisci Elenco Prezzi", key="sostituisci_elenco"):
                for k in ("piano_elenco_bytes", "piano_elenco_nome", "piano_elenco_path"):
                    st.session_state.pop(k, None)
                st.rerun()
            file_elenco = None
        else:
            file_elenco = st.file_uploader(
                "Elenco Prezzi (PDF)",
                type=["pdf"],
                key="upload_elenco",
            )
            if file_elenco:
                st.success(f"✅ {file_elenco.name} — {_get_pagine_cached(file_elenco)} pagine")

        # ── Oneri sicurezza ──────────────────────────────────────────────────
        sic_b_imp = st.session_state.get("piano_sicurezza_bytes")
        sic_n_imp = st.session_state.get("piano_sicurezza_nome", "")
        if sic_b_imp:
            st.success(f"✅ Oneri Sicurezza: **{sic_n_imp}** (importato da Documenti)")
            if st.button("🔄 Sostituisci Sicurezza", key="sostituisci_sicurezza"):
                for k in ("piano_sicurezza_bytes", "piano_sicurezza_nome", "piano_sicurezza_path"):
                    st.session_state.pop(k, None)
                st.rerun()
            file_sicurezza = None
        else:
            file_sicurezza = st.file_uploader(
                "Oneri per la Sicurezza (PDF)",
                type=["pdf"],
                key="upload_sicurezza",
            )
            if file_sicurezza:
                st.success(f"✅ {file_sicurezza.name} — {_get_pagine_cached(file_sicurezza)} pagine")

        ha_cme = bool(files_cme or cme_b_imp)
        ha_sic = bool(file_sicurezza or sic_b_imp)

        if ha_cme or ha_sic:
            if st.button("🔍 Estrai dati da tutti i documenti", type="primary", key="pian_estrai"):
                if not api_key:
                    st.error("API key mancante nella sidebar.")
                else:
                    tutte_cme: list[dict] = []
                    tutte_sic: list[dict] = []
                    ids_cme: list[str] = []

                    if cme_b_imp:
                        with st.spinner(f"Analisi CME importato ({cme_n_imp})…"):
                            tutte_cme.extend(
                                _estrai_voci_cme(_estrai_testo_pdf(cme_b_imp), api_key)
                            )
                        ids_cme.append(f"piano_cme_{cme_n_imp}")

                    if files_cme:
                        ids_cme += [f"{f.name}_{f.size}" for f in files_cme]
                        prog = st.progress(0.0, text="Analisi CME…")
                        for i, f in enumerate(files_cme):
                            prog.progress(
                                (i + 0.5) / len(files_cme),
                                text=f"Analisi {f.name}…",
                            )
                            tutte_cme.extend(
                                _estrai_voci_cme(_estrai_testo_pdf(f.read()), api_key)
                            )
                            f.seek(0)
                        prog.progress(1.0, text="CME completato.")

                    if sic_b_imp:
                        with st.spinner(f"Analisi Sicurezza importata ({sic_n_imp})…"):
                            tutte_sic = _estrai_voci_sicurezza(
                                _estrai_testo_pdf(sic_b_imp), api_key
                            )
                        sic_fid: str | None = f"piano_sicurezza_{sic_n_imp}"
                    elif file_sicurezza:
                        with st.spinner(f"Analisi {file_sicurezza.name}…"):
                            tutte_sic = _estrai_voci_sicurezza(
                                _estrai_testo_pdf(file_sicurezza.read()), api_key
                            )
                            file_sicurezza.seek(0)
                        sic_fid = f"{file_sicurezza.name}_{file_sicurezza.size}"
                    else:
                        sic_fid = None

                    piano["voci_cme"] = tutte_cme
                    piano["voci_sicurezza"] = tutte_sic
                    piano["totale_cme"] = sum(
                        float(v.get("importo_totale") or 0) for v in tutte_cme
                    )
                    piano["totale_sicurezza"] = sum(
                        float(v.get("importo_totale") or 0) for v in tutte_sic
                    )
                    piano["_cme_file_ids"] = ids_cme
                    piano["_sicurezza_file_id"] = sic_fid
                    piano["_associazione_auto_fatta"] = False
                    n_c = len(tutte_cme)
                    n_s = len(tutte_sic)
                    aggiungi_log(
                        "CME estratto",
                        f"{n_c} voci CME + {n_s} voci sicurezza",
                        "Pianificazione",
                    )
                    aggiungi_al_diario(
                        f"CME estratto: {n_c} voci CME + {n_s} voci sicurezza",
                        "🟡 Adempimento Organizzativo",
                    )
                    salva_fn()
                    st.rerun()

        # Stato estratto da sessione
        n_cme_ok = len(piano.get("voci_cme", []))
        n_sic_ok = len(piano.get("voci_sicurezza", []))
        if n_cme_ok or n_sic_ok:
            st.info(f"✅ {n_cme_ok} voci CME + {n_sic_ok} voci sicurezza estratte")

    # ════════════════════════════════════════════════════════════════════════
    # ASSOCIAZIONE FASI → VOCI
    # ════════════════════════════════════════════════════════════════════════
    voci_cme_curr = piano.get("voci_cme", [])
    voci_sic_curr = piano.get("voci_sicurezza", [])
    fasi_curr = piano.get("fasi", [])
    tutte_voci = voci_cme_curr + voci_sic_curr
    n_cme_curr = len(voci_cme_curr)

    if not (fasi_curr and tutte_voci):
        return

    st.divider()
    st.subheader("🔗 Associazione Fasi → Voci")

    if not piano.get("_associazione_auto_fatta"):
        if api_key:
            if st.button(
                "🤖 Associa automaticamente Fasi e Voci",
                type="primary",
                key="pian_auto_assoc",
            ):
                with st.spinner("Claude Haiku associa fasi ↔ voci CME…"):
                    assoc = _associa_fasi_voci(fasi_curr, tutte_voci, api_key)
                if assoc:
                    piano["associazioni"] = assoc
                    piano["_associazione_auto_fatta"] = True
                    aggiungi_log(
                        "Associazione fasi-voci",
                        f"{len(assoc)} fasi associate",
                        "Pianificazione",
                    )
                    salva_fn()
                    st.rerun()
                else:
                    st.warning("Associazione automatica non riuscita. Procedi manualmente.")
        else:
            st.info("Inserisci la API key nella sidebar per l'associazione automatica.")

    # Mappa indice → etichetta (🔴 per voci sicurezza)
    labels_voci = {
        str(i): (
            f"{'🔴 ' if i >= n_cme_curr else ''}"
            f"[{v.get('numero', str(i))}] {v.get('descrizione', '')[:55]}"
        )
        for i, v in enumerate(tutte_voci)
    }
    label_to_id = {v: k for k, v in labels_voci.items()}

    for fi, fase in enumerate(fasi_curr):
        nome_fase = str(fase.get("nome", f"Fase {fi + 1}"))
        ids_assoc = [x for x in piano["associazioni"].get(nome_fase, []) if x in labels_voci]

        voci_f = [tutte_voci[int(i)] for i in ids_assoc if int(i) < len(tutte_voci)]
        budget_cme_f = sum(
            float(v.get("importo_totale") or 0) for v in voci_f
            if v.get("tipo", "cme") != "sicurezza"
        )
        budget_sic_f = sum(
            float(v.get("importo_totale") or 0) for v in voci_f
            if v.get("tipo") == "sicurezza"
        )
        budget_tot_f = budget_cme_f + budget_sic_f

        with st.expander(
            f"📌 {nome_fase} — {fase.get('durata_giorni', '?')} gg "
            f"— S{fase.get('settimana_inizio', '?')}÷S{fase.get('settimana_fine', '?')} "
            f"— € {budget_tot_f:,.0f}",
            expanded=False,
        ):
            default_labels = [labels_voci[i] for i in ids_assoc if i in labels_voci]
            sel = st.multiselect(
                "Voci associate",
                options=list(labels_voci.values()),
                default=default_labels,
                key=f"pian_ms_{fi}",
            )
            nuovi_ids = [label_to_id[l] for l in sel if l in label_to_id]
            if set(nuovi_ids) != set(ids_assoc):
                piano["associazioni"][nome_fase] = nuovi_ids
                salva_fn()
                st.rerun()

            cb1, cb2, cb3 = st.columns(3)
            cb1.metric("💰 Budget CME", f"€ {budget_cme_f:,.2f}")
            cb2.metric("🔴 Budget Sicurezza", f"€ {budget_sic_f:,.2f}")
            cb3.metric("📊 Budget Totale", f"€ {budget_tot_f:,.2f}")

    # ════════════════════════════════════════════════════════════════════════
    # GENERAZIONE EXCEL
    # ════════════════════════════════════════════════════════════════════════
    st.divider()

    col_btn1, col_btn2 = st.columns(2)

    with col_btn1:
        if st.button("📥 Genera Excel Pianificazione", type="primary", key="pian_genera_excel"):
            # Costruisce mappa voce_index → nome_fase
            voce_to_fase: dict[int, str] = {}
            for nome_fase, ids in piano.get("associazioni", {}).items():
                for vid in ids:
                    try:
                        voce_to_fase[int(vid)] = nome_fase
                    except (ValueError, TypeError):
                        pass

            # Aggiunge fase_associata alle voci CME
            voci_cme_ex = []
            for i, v in enumerate(voci_cme_curr):
                v2 = dict(v)
                v2["fase_associata"] = voce_to_fase.get(i, "")
                voci_cme_ex.append(v2)

            # Aggiunge fase_associata e non_ribassabile alle voci sicurezza
            voci_sic_ex = []
            for i, v in enumerate(voci_sic_curr):
                v2 = dict(v)
                v2["fase_associata"] = voce_to_fase.get(n_cme_curr + i, "")
                v2["non_ribassabile"] = True
                voci_sic_ex.append(v2)

            try:
                buf = genera_excel_pianificazione(
                    fasi_curr, voci_cme_ex, voci_sic_ex, nome_progetto
                )
                st.session_state["_excel_pianificazione"] = buf.getvalue()
                aggiungi_log("Excel generato", f"Pianificazione {nome_progetto}", "Pianificazione")
                salva_fn()
            except Exception as e:
                st.error(f"Errore generazione Excel: {e}")

        if "_excel_pianificazione" in st.session_state:
            st.download_button(
                label="📥 Scarica Excel Pianificazione",
                data=st.session_state["_excel_pianificazione"],
                file_name=f"Pianificazione_{nome_progetto}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_pian",
            )

    with col_btn2:
        if st.button("📊 Genera Excel Cruscotto per Fase", key="btn_genera_cruscotto"):
            with st.spinner("Generazione Excel Cruscotto in corso..."):
                excel_bytes = _genera_excel_cruscotto(
                    st.session_state.get("pianificazione", {}),
                    csa_data,
                )
            if excel_bytes:
                st.session_state["_excel_cruscotto"] = excel_bytes
                aggiungi_log(
                    "Excel Cruscotto generato",
                    f"Cruscotto per Fase — {nome_progetto}",
                    "Pianificazione",
                )
                salva_fn()

        if "_excel_cruscotto" in st.session_state:
            comune = csa_data.get("comune", "Cantiere")
            st.download_button(
                label="⬇️ Scarica Excel Cruscotto",
                data=st.session_state["_excel_cruscotto"],
                file_name=f"Cruscotto_Cantiere_{comune}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_cruscotto",
            )
