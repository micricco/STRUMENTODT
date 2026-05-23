"""
rubrica_tab.py — Rubrica Contatti cantiere
Inserimento manuale soggetti per ruolo: SA, Impresa, Sub, Fornitori,
Professionisti, Enti. Telefono/email cliccabili, ricerca, export Excel.
"""

import io
import uuid

import streamlit as st


# ── Ruoli predefiniti ──────────────────────────────────────────────────────────

_RUOLI_APPALTO = {
    "stazione_appaltante": {
        "label": "Stazione Appaltante",
        "icona": "🏛️",
        "sotto_ruoli": [
            "RUP (Responsabile Unico del Procedimento)",
            "DL (Direttore dei Lavori)",
            "CSE (Coordinatore Sicurezza in Esecuzione)",
            "Collaudatore",
            "Responsabile Finanziario",
            "Ufficio Tecnico",
        ],
    },
    "impresa": {
        "label": "Impresa Appaltatrice",
        "icona": "🏗️",
        "sotto_ruoli": [
            "PM (Project Manager)",
            "DTC (Direttore Tecnico di Cantiere)",
            "Capocantiere",
            "Responsabile Sicurezza",
            "Amministrazione",
            "Direzione",
        ],
    },
    "subappaltatori": {
        "label": "Subappaltatori",
        "icona": "🤝",
        "sotto_ruoli": [
            "Subappaltatore",
            "Subaffidatario",
            "Cottimista",
        ],
    },
    "fornitori": {
        "label": "Fornitori",
        "icona": "📦",
        "sotto_ruoli": [
            "Fornitore Materiali",
            "Noleggiatore",
            "Trasportatore",
        ],
    },
    "professionisti": {
        "label": "Professionisti",
        "icona": "👷",
        "sotto_ruoli": [
            "Progettista",
            "Geologo",
            "Coordinatore Sicurezza in Progettazione (CSP)",
            "Collaudatore Statico",
            "Ispettore di Cantiere",
        ],
    },
    "enti": {
        "label": "Enti e Uffici",
        "icona": "🏢",
        "sotto_ruoli": [
            "Comune — Ufficio Tecnico",
            "ASL — Ispettorato",
            "Vigili del Fuoco",
            "INAIL",
            "Ispettorato del Lavoro",
            "Gestore Reti (Enel/Gas/Acqua)",
        ],
    },
}


# ── Export Excel ───────────────────────────────────────────────────────────────

def _genera_excel_rubrica(rubrica: dict) -> bytes | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Rubrica Cantiere"

        headers = ["CATEGORIA", "RUOLO", "NOME", "TELEFONO", "EMAIL", "NOTE"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E5099")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 25

        row = 2
        for cat_key, info in _RUOLI_APPALTO.items():
            for c in rubrica.get(cat_key, []):
                ws.cell(row=row, column=1, value=info["label"])
                ws.cell(row=row, column=2, value=c.get("sotto_ruolo", ""))
                ws.cell(row=row, column=3, value=c.get("nome", ""))
                ws.cell(row=row, column=4, value=c.get("telefono", ""))
                ws.cell(row=row, column=5, value=c.get("email", ""))
                ws.cell(row=row, column=6, value=c.get("note", ""))
                row += 1

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return None


# ── Render principale ──────────────────────────────────────────────────────────

def render_rubrica_tab(salva_fn) -> None:
    st.subheader("📒 Rubrica Contatti")
    st.caption(
        "Inserisci i recapiti di tutti i soggetti coinvolti nell'appalto. "
        "Puoi chiamare o scrivere email direttamente da qui."
    )

    rubrica: dict = st.session_state.get("rubrica_contatti", {})

    # ── Form aggiunta contatto ─────────────────────────────────────────────────
    with st.expander("➕ Aggiungi Contatto", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            categoria = st.selectbox(
                "Categoria",
                list(_RUOLI_APPALTO.keys()),
                format_func=lambda k: f"{_RUOLI_APPALTO[k]['icona']} {_RUOLI_APPALTO[k]['label']}",
                key="rb_categoria",
            )
            sotto_ruolo = st.selectbox(
                "Ruolo specifico",
                _RUOLI_APPALTO[categoria]["sotto_ruoli"] + ["Altro"],
                key="rb_sotto_ruolo",
            )
            nome = st.text_input(
                "Nome e Cognome / Ragione Sociale",
                key="rb_nome",
                placeholder="Es: Mario Rossi",
            )
        with col2:
            telefono = st.text_input(
                "📞 Telefono",
                key="rb_telefono",
                placeholder="Es: +39 333 1234567",
            )
            email = st.text_input(
                "📧 Email",
                key="rb_email",
                placeholder="Es: mario.rossi@comune.it",
            )
            note = st.text_input(
                "Note",
                key="rb_note",
                placeholder="Es: disponibile lun-ven 9-17",
            )

        if st.button("💾 Salva Contatto", key="btn_salva_contatto"):
            if not nome.strip():
                st.error("❌ Inserisci almeno il nome")
            else:
                contatto_id = f"cont_{uuid.uuid4().hex[:8]}"
                rubrica.setdefault(categoria, []).append({
                    "id": contatto_id,
                    "sotto_ruolo": sotto_ruolo,
                    "nome": nome.strip(),
                    "telefono": telefono.strip(),
                    "email": email.strip(),
                    "note": note.strip(),
                })
                st.session_state["rubrica_contatti"] = rubrica
                salva_fn()
                st.success(f"✅ Contatto '{nome.strip()}' salvato")
                st.rerun()

    st.divider()

    # ── Ricerca rapida ─────────────────────────────────────────────────────────
    ricerca = st.text_input(
        "🔍 Cerca contatto",
        key="rb_ricerca",
        placeholder="Nome, ruolo, email…",
    )

    # ── Lista contatti per categoria ───────────────────────────────────────────
    totale_contatti = sum(len(v) for v in rubrica.values())

    if totale_contatti == 0:
        st.info("Nessun contatto inserito. Clicca '➕ Aggiungi Contatto' per iniziare.")
        return

    for cat_key, info in _RUOLI_APPALTO.items():
        contatti_cat = rubrica.get(cat_key, [])

        if ricerca.strip():
            q = ricerca.lower()
            contatti_cat = [
                c for c in contatti_cat
                if q in c.get("nome", "").lower()
                or q in c.get("sotto_ruolo", "").lower()
                or q in c.get("email", "").lower()
                or q in c.get("telefono", "").lower()
            ]

        if not contatti_cat:
            continue

        st.subheader(f"{info['icona']} {info['label']} ({len(contatti_cat)})")

        for contatto in contatti_cat:
            cid = contatto.get("id", "")
            with st.container(border=True):
                col_a, col_b, col_c, col_d = st.columns([3, 2, 2, 1])

                with col_a:
                    st.markdown(f"**{contatto.get('nome', '—')}**")
                    st.caption(contatto.get("sotto_ruolo", "—"))
                    if contatto.get("note"):
                        st.caption(f"📝 {contatto['note']}")

                with col_b:
                    tel = contatto.get("telefono", "")
                    if tel:
                        tel_clean = tel.replace(" ", "")
                        st.markdown(f"📞 [{tel}](tel:{tel_clean})")
                    else:
                        st.caption("📞 —")

                with col_c:
                    em = contatto.get("email", "")
                    if em:
                        st.markdown(f"📧 [{em}](mailto:{em})")
                    else:
                        st.caption("📧 —")

                with col_d:
                    if st.button("🗑️", key=f"del_cont_{cid}", help="Elimina contatto"):
                        rubrica[cat_key] = [
                            c for c in rubrica.get(cat_key, [])
                            if c.get("id") != cid
                        ]
                        st.session_state["rubrica_contatti"] = rubrica
                        salva_fn()
                        st.rerun()

    # ── Export rubrica ─────────────────────────────────────────────────────────
    st.divider()
    if st.button("📊 Esporta Rubrica Excel", key="btn_export_rubrica"):
        excel = _genera_excel_rubrica(rubrica)
        if excel:
            st.download_button(
                "⬇️ Scarica Excel",
                data=excel,
                file_name="Rubrica_Cantiere.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_rubrica",
            )
        else:
            st.error("Errore generazione Excel — installa openpyxl")
