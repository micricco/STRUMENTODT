"""
app.py — Strumento DTC (Direttore Tecnico di Cantiere)
Analisi CSA e gestione appalti pubblici italiani (D.Lgs. 36/2023, DM 49/2018)
"""

import json
import math
import os
import pathlib
import re
import smtplib
import time
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

import anthropic
import pandas as pd
import streamlit as st

from modules.calendar_manager import Sospensione, calcola_calendario
from modules.csa_analyzer import (
    MODEL_FAST,
    _COSTO_INPUT_HAIKU,
    _COSTO_OUTPUT_HAIKU,
    _estrai_pagine_rilevanti,
    analyze_csa,
    conta_pagine_pdf,
    conta_token_api,
    stima_token_pdf,
)
from modules.contract_analyzer import generate_document
from modules.doc_viewer import render_doc_buttons
from modules.log_manager import aggiungi_log, render_log_tab
from modules.operatori_tab import render_durc_semaphore, render_operatori_tab
from modules.penalties import calcola_penale_cumulativa, simula_revisione_prezzi
from modules.registri_tab import render_registri_tab, _render_contabilita_sal
from modules.pianificazione_tab import render_pianificazione_tab

# ── Costanti ───────────────────────────────────────────────────────────────────
RESULTS_DIR = pathlib.Path("results")
FONTS_DIR = pathlib.Path("fonts")
_FONT_REGULAR = str(FONTS_DIR / "DejaVuSans.ttf")
_FONT_BOLD = str(FONTS_DIR / "DejaVuSans-Bold.ttf")

_CHECKLIST_CATEGORIE = {
    "Prime settimane": "checklist_prime_settimane",
    "Sicurezza": "checklist_sicurezza",
    "Assicurative": "checklist_assicurative",
}

_STATI_CHECKLIST = ["Da fare", "In corso", "Completato", "In ritardo"]
_STATI_COLORI = {
    "Da fare": "⚪",
    "In corso": "🔵",
    "Completato": "🟢",
    "In ritardo": "🔴",
}

_TIPI_FILE_DOC = ["pdf", "docx", "doc", "xlsx", "jpg", "png"]

# ── Streamlit config ───────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Strumento DTC — Appalti Pubblici",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: parse importo
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_importo(valore) -> float:
    if valore is None:
        return 0.0
    if isinstance(valore, (int, float)):
        return float(valore)
    s = str(valore).strip()
    if not s or s.lower() in ("null", "none", "—", "-"):
        return 0.0
    # Formato italiano: 1.350.000,00
    s_clean = re.sub(r"[€\s]", "", s)
    if "," in s_clean and "." in s_clean:
        s_clean = s_clean.replace(".", "").replace(",", ".")
    elif "," in s_clean:
        s_clean = s_clean.replace(",", ".")
    try:
        return float(s_clean)
    except ValueError:
        return _parse_importo_italiano(s)


def _parse_importo_italiano(testo: str) -> float:
    mapping = {
        "un milione": 1_000_000, "due milioni": 2_000_000,
        "tre milioni": 3_000_000, "quattro milioni": 4_000_000,
        "cinque milioni": 5_000_000,
        "centomila": 100_000, "duecentomila": 200_000,
        "trecentomila": 300_000, "quattrocentomila": 400_000,
        "cinquecentomila": 500_000, "seicentomila": 600_000,
        "settecentomila": 700_000, "ottocentomila": 800_000,
        "novecentomila": 900_000,
        "cinquantamila": 50_000, "centomila": 100_000,
    }
    t = testo.lower()
    for k, v in mapping.items():
        if k in t:
            return float(v)
    nums = re.findall(r"[\d.,]+", testo)
    for n in nums:
        try:
            return _parse_importo(n)
        except Exception:
            pass
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: rate limit / retry
# ═══════════════════════════════════════════════════════════════════════════════

def _chiama_con_rate_limit(fn, *args, max_retry: int = 2, wait_sec: int = 60, **kwargs):
    for tentativo in range(max_retry + 1):
        try:
            return fn(*args, **kwargs)
        except anthropic.RateLimitError:
            if tentativo == max_retry:
                raise
            placeholder = st.empty()
            for sec in range(wait_sec, 0, -1):
                placeholder.warning(
                    f"⏳ Rate limit API (429) — attesa {sec}s (tentativo {tentativo + 1}/{max_retry})…"
                )
                time.sleep(1)
            placeholder.empty()


# ═══════════════════════════════════════════════════════════════════════════════
# SALVATAGGIO / CARICAMENTO ANALISI
# ═══════════════════════════════════════════════════════════════════════════════

def _slug_nome(nome_pdf: str) -> str:
    slug = re.sub(r"[^\w\-]", "_", pathlib.Path(nome_pdf).stem)
    return slug[:60]


def _salva_analisi() -> None:
    csa_data = st.session_state.get("csa_data")
    if not csa_data:
        return
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    nome_pdf = st.session_state.get("_pdf_nome", "analisi")
    slug = _slug_nome(nome_pdf)
    percorso = RESULTS_DIR / f"{slug}.json"

    payload = dict(csa_data)
    # Merge CIG/CUP inseriti manualmente
    if st.session_state.get("cig_manuale"):
        payload["cig"] = st.session_state["cig_manuale"]
    if st.session_state.get("cup_manuale"):
        payload["cup"] = st.session_state["cup_manuale"]
    payload["_ribasso_pct"] = float(st.session_state.get("ribasso_pct", 0.0))
    payload["_checklist_stato"] = st.session_state.get("checklist_stato", {})
    payload["_log_attivita"] = st.session_state.get("log_attivita", [])
    payload["_operatori_economici"] = st.session_state.get("operatori_economici", {})
    payload["_registri"] = st.session_state.get("registri", {})

    dc = st.session_state.get("data_consegna_cantiere")
    payload["_data_consegna_cantiere"] = dc.isoformat() if dc else None

    payload["_doc_elaborati"] = st.session_state.get("doc_elaborati", {})
    payload["_varianti_proroghe"] = st.session_state.get("_varianti_proroghe", [])

    # pianificazione
    piano = st.session_state.get("pianificazione")
    if piano:
        payload["_pianificazione"] = piano

    try:
        with open(percorso, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        st.warning(f"Errore salvataggio: {e}")


def _salva_stato_cantiere() -> None:
    _salva_analisi()


def _lista_analisi_salvate() -> list[pathlib.Path]:
    if not RESULTS_DIR.exists():
        return []
    files = sorted(RESULTS_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files


def _carica_analisi(percorso: pathlib.Path) -> None:
    try:
        with open(percorso, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception as e:
        st.error(f"Errore caricamento: {e}")
        return

    # Estrai stato esteso
    ribasso = payload.pop("_ribasso_pct", 0.0)
    st.session_state.checklist_stato = payload.pop("_checklist_stato", {})
    st.session_state.log_attivita = payload.pop("_log_attivita", [])
    st.session_state.operatori_economici = payload.pop("_operatori_economici", {})
    st.session_state.registri = payload.pop("_registri", {})

    dc_str = payload.pop("_data_consegna_cantiere", None)
    if dc_str:
        try:
            st.session_state.data_consegna_cantiere = date.fromisoformat(dc_str)
        except Exception:
            st.session_state.data_consegna_cantiere = None
    else:
        st.session_state.data_consegna_cantiere = None

    st.session_state.doc_elaborati = payload.pop("_doc_elaborati", {})
    st.session_state._varianti_proroghe = payload.pop("_varianti_proroghe", [])

    piano = payload.pop("_pianificazione", None)
    if piano:
        st.session_state.pianificazione = piano

    st.session_state.csa_data = payload
    # Ripristina CIG/CUP manuale in session_state
    if payload.get("cig") and payload["cig"] != "—":
        st.session_state["cig_manuale"] = payload["cig"]
    if payload.get("cup") and payload["cup"] != "—":
        st.session_state["cup_manuale"] = payload["cup"]
    st.session_state._file_id = f"__saved__{percorso.name}"
    st.session_state._pdf_nome = percorso.stem
    st.session_state._demo_active = False
    st.session_state.pdf_bytes = None

    # Staging ribasso (per evitare StreamlitAPIException)
    st.session_state._ribasso_pendente = float(ribasso)

    # Reset altri stati
    st.session_state.sospensioni = []
    st.session_state._sosp_counter = 0
    st.session_state.coords = None
    st.session_state.suppliers = None

    aggiungi_log("Analisi caricata", percorso.name, tab="Sidebar")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _genera_excel(csa_data: dict, importo_netto: float) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill

    wb = Workbook()
    _BLUE = "1F4E79"
    _WHITE = "FFFFFF"
    _LGRAY = "F2F2F2"

    def _hdr(ws, row, col, val, bold=True, bg=_BLUE, fg=_WHITE, size=10):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        return c

    # ── Foglio 1: Sintesi CSA ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sintesi CSA"
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 50

    _hdr(ws1, 1, 1, "SINTESI CSA — DATI PRINCIPALI", size=12)
    ws1.merge_cells("A1:B1")

    campi = [
        ("Stazione Appaltante", csa_data.get("stazione_appaltante", "")),
        ("Tipo Lavori", csa_data.get("tipo_lavori", "")),
        ("Indirizzo Cantiere", csa_data.get("indirizzo_cantiere", "")),
        ("Comune", csa_data.get("comune", "")),
        ("Provincia", csa_data.get("provincia", "")),
        ("CIG", csa_data.get("cig", "")),
        ("CUP", csa_data.get("cup", "")),
        ("Importo Lavori (lordo)", csa_data.get("importo_lavori", "")),
        ("Importo Netto (dopo ribasso)", f"€ {importo_netto:,.2f}"),
        ("Durata (giorni)", csa_data.get("durata_lavori_giorni", "")),
        ("Tipo Contratto", csa_data.get("tipo_contratto", "")),
        ("Prezzario", f"{csa_data.get('prezzario_nome', '')} {csa_data.get('prezzario_anno', '')}"),
        ("SAL Tipo", csa_data.get("sal_tipo", "")),
        ("SAL Intervallo (gg)", csa_data.get("sal_intervallo_giorni", "")),
        ("SAL Importo Minimo (€)", csa_data.get("sal_importo_minimo_euro", "")),
        ("Penale Giornaliera (‰)", csa_data.get("penale_giornaliera_permille", "")),
        ("Penale Massima (%)", csa_data.get("penale_massima_percentuale", "")),
        ("Riserve Iscrizione (gg)", csa_data.get("riserve_iscrizione_giorni", "")),
        ("Riserve Quantificazione (gg)", csa_data.get("riserve_quantificazione_giorni", "")),
        ("Collaudo (gg)", csa_data.get("collaudo_giorni", "")),
        ("Oneri Sicurezza (€)", csa_data.get("importo_oneri_sicurezza", "")),
        ("Subappalto max %", csa_data.get("subappalto_percentuale_massima", "")),
    ]
    for r, (label, val) in enumerate(campi, 2):
        ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=r, column=2, value=str(val) if val is not None else "")

    # SOA
    row_soa = len(campi) + 3
    _hdr(ws1, row_soa, 1, "CATEGORIE SOA")
    for ci, h in enumerate(["Codice", "Descrizione", "Classifica", "Prevalente"], 2):
        _hdr(ws1, row_soa, ci, h)
    ws1.merge_cells(f"A{row_soa}:A{row_soa}")
    for i, soa in enumerate(csa_data.get("categorie_soa", []), 1):
        r = row_soa + i
        ws1.cell(row=r, column=1, value=soa.get("codice", ""))
        ws1.cell(row=r, column=2, value=soa.get("descrizione_categoria", ""))
        ws1.cell(row=r, column=3, value=soa.get("classifica", ""))
        ws1.cell(row=r, column=4, value="Sì" if soa.get("prevalente") else "No")
        ws1.cell(row=r, column=5, value=soa.get("motivazione", ""))

    # ── Foglio 2: Scadenze + Obblighi ─────────────────────────────────────────
    ws2 = wb.create_sheet("Scadenze e Obblighi")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 15

    _hdr(ws2, 1, 1, "OBBLIGHI APPALTATORE")
    ws2.merge_cells("A1:C1")
    for i, ob in enumerate(csa_data.get("obblighi_appaltatore", []), 2):
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=ob)

    offset = len(csa_data.get("obblighi_appaltatore", [])) + 3
    _hdr(ws2, offset, 1, "OBBLIGHI STAZIONE APPALTANTE")
    ws2.merge_cells(f"A{offset}:C{offset}")
    for i, ob in enumerate(csa_data.get("obblighi_stazione_appaltante", []), 1):
        ws2.cell(row=offset + i, column=1, value=i)
        ws2.cell(row=offset + i, column=2, value=ob)

    # ── Foglio 3: Checklist ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Checklist")
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 15
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 15

    for ci, h in enumerate(["#", "Attività", "Categoria", "Priorità", "Termine (gg)"], 1):
        _hdr(ws3, 1, ci, h)

    r3 = 2
    checklist_stato = st.session_state.get("checklist_stato", {})
    for cat_label, cat_key in _CHECKLIST_CATEGORIE.items():
        voci = csa_data.get(cat_key, [])
        for idx, v in enumerate(voci):
            stato_voce = (checklist_stato.get(cat_key, {}).get(str(idx), {}) or {}).get("stato", "Da fare")
            ws3.cell(row=r3, column=1, value=idx + 1)
            ws3.cell(row=r3, column=2, value=v.get("attivita", ""))
            ws3.cell(row=r3, column=3, value=cat_label)
            ws3.cell(row=r3, column=4, value=v.get("priorita", ""))
            ws3.cell(row=r3, column=5, value=v.get("termine_giorni", ""))
            ws3.cell(row=r3, column=6, value=stato_voce)
            r3 += 1

    # ── Foglio 4: Calendario ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Calendario")
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 50
    ws4.column_dimensions["D"].width = 12

    for ci, h in enumerate(["Data", "Tipo", "Descrizione", "Giorno Contr."], 1):
        _hdr(ws4, 1, ci, h)

    dc = st.session_state.get("data_consegna_cantiere") or date.today()
    durata = int(csa_data.get("durata_lavori_giorni") or 180)
    ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
    importo_base = _parse_importo(csa_data.get("importo_lavori"))
    importo_netto_cal = importo_base * (1 - ribasso_pct / 100)

    vp = st.session_state.get("_varianti_proroghe", [])
    giorni_varianti = sum(int(v.get("giorni", 0) or 0) for v in vp)
    durata_eff = durata + giorni_varianti

    sal_tipo = csa_data.get("sal_tipo", "tempo")
    sal_gg = csa_data.get("sal_intervallo_giorni")
    sal_imp = csa_data.get("sal_importo_minimo_euro") or (
        importo_netto_cal * (csa_data.get("sal_percentuale_minima") or 0) / 100
        if csa_data.get("sal_percentuale_minima") else None
    )

    cal = calcola_calendario(
        data_consegna=dc,
        durata_giorni=durata_eff,
        importo_contratto=importo_netto_cal,
        sal_tipo=sal_tipo,
        sal_intervallo_giorni=sal_gg,
        sal_importo_soglia=sal_imp,
        sospensioni=st.session_state.get("sospensioni", []),
    )

    for r4, ev in enumerate(cal["eventi"], 2):
        ws4.cell(row=r4, column=1, value=ev["data"].isoformat())
        ws4.cell(row=r4, column=2, value=ev["tipo"])
        ws4.cell(row=r4, column=3, value=ev["descrizione"])
        ws4.cell(row=r4, column=4, value=ev["giorno_contratto"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICHE EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

def _invia_email_scadenze(
    smtp_server: str, smtp_port: int, mittente: str, password: str,
    destinatario: str, csa_data: dict, giorni_soglia: int
) -> tuple[bool, str]:
    dc = st.session_state.get("data_consegna_cantiere") or date.today()
    durata = int(csa_data.get("durata_lavori_giorni") or 180)
    importo_base = _parse_importo(csa_data.get("importo_lavori"))
    ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
    importo_netto = importo_base * (1 - ribasso_pct / 100)

    cal = calcola_calendario(
        data_consegna=dc,
        durata_giorni=durata,
        importo_contratto=importo_netto,
        sal_tipo=csa_data.get("sal_tipo", "tempo"),
        sal_intervallo_giorni=csa_data.get("sal_intervallo_giorni"),
        sal_importo_soglia=csa_data.get("sal_importo_minimo_euro"),
        sospensioni=st.session_state.get("sospensioni", []),
    )

    oggi = date.today()
    eventi_critici = [
        ev for ev in cal["eventi"]
        if ev.get("critico") and 0 <= (ev["data"] - oggi).days <= giorni_soglia
    ]

    if not eventi_critici:
        return True, "Nessuna scadenza critica entro il periodo selezionato."

    righe_html = "".join(
        f"<tr><td>{ev['data'].strftime('%d/%m/%Y')}</td>"
        f"<td>{ev['tipo']}</td>"
        f"<td>{ev['descrizione']}</td>"
        f"<td style='color:{'red' if (ev['data']-oggi).days<=7 else 'orange'}'>"
        f"{(ev['data']-oggi).days} gg</td></tr>"
        for ev in eventi_critici
    )

    corpo_html = f"""
<html><body>
<h2>Scadenze DTC — {csa_data.get('tipo_lavori','')}</h2>
<p>Cantiere: {csa_data.get('comune','')} ({csa_data.get('provincia','')})</p>
<table border="1" cellpadding="5" cellspacing="0" style="border-collapse:collapse">
<tr style="background:#1F4E79;color:#fff">
<th>Data</th><th>Tipo</th><th>Descrizione</th><th>Giorni rimanenti</th>
</tr>
{righe_html}
</table>
<p style="font-size:11px;color:#888">Strumento DTC — D.Lgs. 36/2023</p>
</body></html>"""

    try:
        msg = MIMEMultipart("alternative")
        oggetto_ascii = f"DTC Scadenze: {csa_data.get('comune','')} - {len(eventi_critici)} eventi"
        msg["Subject"] = oggetto_ascii.encode("ascii", errors="replace").decode()
        msg["From"] = mittente
        msg["To"] = destinatario
        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(mittente, password)
            server.send_message(msg)
        return True, f"Email inviata con {len(eventi_critici)} scadenze critiche."
    except Exception as e:
        return False, str(e)


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════════════════════

def _init_session_state() -> None:
    defaults = {
        "_file_id": None,
        "_demo_active": False,
        "_pdf_nome": "",
        "pdf_bytes": None,
        "csa_data": None,
        "coords": None,
        "suppliers": None,
        "sospensioni": [],
        "_sosp_counter": 0,
        "checklist_stato": {},
        "data_consegna_cantiere": None,
        "operatori_economici": {},
        "log_attivita": [],
        "doc_elaborati": {"grafico": [], "tecnico": [], "amministrativo": []},
        "_varianti_proroghe": [],
        "registri": {
            "riserve": [], "verbali": [], "non_conformita": [],
            "ordini_servizio": [], "contabilita_sal": []
        },
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ═══════════════════════════════════════════════════════════════════════════════
# CALLBACK ribasso
# ═══════════════════════════════════════════════════════════════════════════════

def _on_ribasso_change() -> None:
    aggiungi_log(
        "Ribasso d'asta modificato",
        f"{st.session_state.get('ribasso_pct', 0.0):.2f}%",
        tab="Sidebar",
    )
    _salva_analisi()


# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════

def _render_sidebar() -> str:
    with st.sidebar:
        st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/0/03/Flag_of_Italy.svg/22px-Flag_of_Italy.svg.png", width=22)
        st.title("🏗️ Strumento DTC")
        st.caption("Appalti pubblici — D.Lgs. 36/2023")
        st.divider()

        # ── API Key ────────────────────────────────────────────────────────────
        api_key = st.text_input(
            "Claude API Key",
            type="password",
            placeholder="sk-ant-…",
            key="_api_key_input",
            help="Anthropic API Key — mai salvata su disco",
        )

        st.divider()

        # ── Ribasso d'asta (staging pattern) ──────────────────────────────────
        if "_ribasso_pendente" in st.session_state:
            pendente = st.session_state.pop("_ribasso_pendente")
            st.session_state["ribasso_pct"] = float(pendente)

        st.number_input(
            "Ribasso d'asta (%)",
            min_value=0.0,
            max_value=99.0,
            step=0.01,
            format="%.2f",
            key="ribasso_pct",
            on_change=_on_ribasso_change,
            help="Il ribasso viene applicato a Calendario, Penali e Revisione Prezzi",
        )

        # ── Modalità demo ──────────────────────────────────────────────────────
        st.divider()
        col_demo, col_reset = st.columns(2)
        with col_demo:
            if st.button("🎮 Demo", use_container_width=True, help="Dati fittizi Bergamo BG", key="sb_demo_btn"):
                _attiva_demo()
        with col_reset:
            if st.button("🔄 Reset", use_container_width=True, help="Pulisci sessione", key="sb_reset_btn"):
                _reset_sessione()

        # ── Upload PDF ─────────────────────────────────────────────────────────
        st.divider()
        st.markdown("### 📤 Carica CSA")
        uploaded = st.file_uploader(
            "Seleziona PDF del Capitolato Speciale d'Appalto",
            type=["pdf"],
            key="pdf_uploader",
            label_visibility="collapsed",
        )

        if uploaded is not None:
            file_id_nuovo = f"{uploaded.name}_{uploaded.size}"
            if file_id_nuovo != st.session_state.get("_file_id"):
                st.session_state._file_id = file_id_nuovo
                st.session_state._pdf_nome = uploaded.name
                st.session_state.pdf_bytes = uploaded.read()
                st.session_state.csa_data = None
                st.session_state._demo_active = False
                st.session_state.sospensioni = []
                st.session_state._sosp_counter = 0
                st.session_state.checklist_stato = {}
                st.session_state.doc_elaborati = {"grafico": [], "tecnico": [], "amministrativo": []}
                st.session_state._varianti_proroghe = []
                st.rerun()

        # ── Analisi CSA ────────────────────────────────────────────────────────
        if st.session_state.get("pdf_bytes") and not st.session_state.get("csa_data"):
            pdf_bytes = st.session_state.pdf_bytes
            n_pagine = conta_pagine_pdf(pdf_bytes)
            stima_tok = stima_token_pdf(pdf_bytes)
            costo_est = stima_tok * _COSTO_INPUT_HAIKU
            st.info(
                f"📄 **{st.session_state._pdf_nome}**\n\n"
                f"Pagine: {n_pagine} — Token stimati: ~{stima_tok:,}\n\n"
                f"Costo stimato: ~${costo_est:.3f}"
            )
            if api_key:
                if st.button("🔍 Avvia analisi CSA", type="primary", use_container_width=True, key="sb_avvia_btn"):
                    _esegui_analisi(pdf_bytes, api_key)
            else:
                st.warning("Inserisci la API Key per avviare l'analisi.")

        # ── Analisi salvate ────────────────────────────────────────────────────
        analisi_salvate = _lista_analisi_salvate()
        file_id_attivo = st.session_state.get("_file_id", "")

        with st.expander("📂 Analisi salvate", expanded=bool(analisi_salvate)):
            if not analisi_salvate:
                st.caption("Nessuna analisi salvata. Dopo la prima analisi apparirà qui.")
            else:
                for percorso in analisi_salvate:
                    nome_j = percorso.stem
                    data_mod = datetime.fromtimestamp(percorso.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
                    is_attiva = file_id_attivo == f"__saved__{percorso.name}"
                    badge = " ✅" if is_attiva else ""
                    col_car, col_del = st.columns([4, 1])
                    with col_car:
                        if st.button(
                            f"{nome_j[:25]}{badge}\n{data_mod}",
                            key=f"load_{percorso.name}",
                            use_container_width=True,
                        ):
                            _carica_analisi(percorso)
                            st.rerun()
                    with col_del:
                        if st.button("🗑️", key=f"del_{percorso.name}", help="Elimina"):
                            percorso.unlink(missing_ok=True)
                            if is_attiva:
                                _reset_sessione()
                            st.rerun()

        # ── Export Excel ───────────────────────────────────────────────────────
        csa_data = st.session_state.get("csa_data")
        if csa_data:
            st.divider()
            with st.expander("📥 Esporta"):
                ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
                importo_base = _parse_importo(csa_data.get("importo_lavori"))
                importo_netto = importo_base * (1 - ribasso_pct / 100)

                if st.button("Genera Excel multi-foglio", use_container_width=True, key="sb_genera_excel_btn"):
                    try:
                        xls_bytes = _genera_excel(csa_data, importo_netto)
                        nome_xls = f"DTC_{_slug_nome(st.session_state._pdf_nome)}.xlsx"
                        st.download_button(
                            "⬇️ Scarica Excel",
                            data=xls_bytes,
                            file_name=nome_xls,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="sb_dl_excel_btn",
                        )
                        aggiungi_log("Export Excel", nome_xls, tab="Sidebar")
                    except Exception as e:
                        st.error(f"Errore Excel: {e}")

            # ── Notifiche email ────────────────────────────────────────────────
            with st.expander("📧 Notifiche email scadenze"):
                giorni_soglia = st.slider(
                    "Scadenze entro (giorni)", 7, 90, 30, key="email_giorni"
                )
                smtp_srv = st.text_input("Server SMTP", value="smtp.gmail.com", key="email_smtp")
                smtp_prt = st.number_input("Porta", value=587, step=1, key="email_porta")
                mitt = st.text_input("Email mittente", key="email_mitt")
                pwd = st.text_input("Password app", type="password", key="email_pwd")
                dest = st.text_input("Destinatario", key="email_dest")
                if st.button("Invia notifica", use_container_width=True, key="email_invia"):
                    ok, msg = _invia_email_scadenze(smtp_srv, smtp_prt, mitt, pwd, dest, csa_data, giorni_soglia)
                    (st.success if ok else st.error)(msg)

    return api_key


# ═══════════════════════════════════════════════════════════════════════════════
# DEMO / RESET
# ═══════════════════════════════════════════════════════════════════════════════

def _attiva_demo() -> None:
    from modules.demo_data import DEMO_CSA_DATA, DEMO_COORDS, DEMO_SUPPLIERS, DEMO_DOCUMENTS, DEMO_REGISTRI
    st.session_state.csa_data = DEMO_CSA_DATA.copy()
    st.session_state.registri = {k: list(v) for k, v in DEMO_REGISTRI.items()}
    st.session_state._demo_active = True
    st.session_state._file_id = "__demo__"
    st.session_state._pdf_nome = "demo_bergamo"
    st.session_state.pdf_bytes = None
    st.session_state.coords = DEMO_COORDS
    st.session_state.suppliers = DEMO_SUPPLIERS
    st.session_state.sospensioni = []
    st.session_state._sosp_counter = 0
    st.session_state.checklist_stato = {}
    st.session_state.doc_elaborati = {"grafico": [], "tecnico": [], "amministrativo": []}
    st.session_state._varianti_proroghe = []
    for doc_type, testo in DEMO_DOCUMENTS.items():
        st.session_state[f"_doc_{doc_type}"] = testo
    aggiungi_log("Modalità demo attivata", "Cantiere Bergamo BG", tab="Sidebar")
    st.rerun()


def _reset_sessione() -> None:
    chiavi = [
        "_file_id", "_demo_active", "_pdf_nome", "pdf_bytes", "csa_data",
        "coords", "suppliers", "sospensioni", "_sosp_counter",
        "checklist_stato", "data_consegna_cantiere", "operatori_economici",
        "log_attivita", "doc_elaborati", "_varianti_proroghe", "registri",
        "_smart_extract_stats", "_durc_alerts_logged",
        "_doc_riserva", "_doc_verbale_consegna", "_doc_proroga", "_doc_contestazione",
        "pianificazione",
    ]
    for k in chiavi:
        st.session_state.pop(k, None)
    st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# ANALISI CSA
# ═══════════════════════════════════════════════════════════════════════════════

def _esegui_analisi(pdf_bytes: bytes, api_key: str) -> None:
    barra = st.progress(0.0, text="Smart Extract — filtro pagine rilevanti…")
    try:
        testo_filtrato, stats = _estrai_pagine_rilevanti(pdf_bytes)
        st.session_state._smart_extract_stats = stats
        barra.progress(0.30, text="Conteggio token…")

        n_token = conta_token_api(testo_filtrato, api_key)
        costo_est = n_token * _COSTO_INPUT_HAIKU
        barra.progress(0.50, text=f"Token: {n_token:,} (~${costo_est:.3f}) — Analisi in corso…")

        csa_data = _chiama_con_rate_limit(
            analyze_csa, testo_filtrato, api_key
        )
        barra.progress(0.95, text="Salvataggio…")
        st.session_state.csa_data = csa_data
        st.session_state._demo_active = False
        _salva_analisi()
        aggiungi_log(
            "CSA analizzato",
            f"{st.session_state._pdf_nome} — {stats.get('n_pagine_estratte', 0)}/{stats.get('n_pagine_totali', 0)} pagine",
            tab="Sidebar",
        )
        barra.progress(1.0, text="Analisi completata.")
        st.success("✅ Analisi completata. Esplora i tab.")
        st.rerun()
    except Exception as e:
        barra.empty()
        st.error(f"Errore analisi: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 0 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def _render_dashboard(csa_data: dict, details: dict, importo_netto: float) -> None:
    st.header("🏠 Dashboard Cantiere")

    tipo = csa_data.get("tipo_lavori", "—")
    comune = csa_data.get("comune", "")
    prov = csa_data.get("provincia", "")
    sa = csa_data.get("stazione_appaltante", "—")

    st.markdown(f"**{tipo}**  \n{sa}  \n{comune} ({prov})")

    cig = csa_data.get("cig", "") or ""
    cup = csa_data.get("cup", "") or ""
    col_cig1, col_cig2 = st.columns(2)
    with col_cig1:
        if not cig or cig == "—":
            cig_input = st.text_input(
                "CIG (non rilevato — inserisci manualmente)",
                value=st.session_state.get("cig_manuale", ""),
                placeholder="Es: 9999999999",
                key="input_cig_manuale",
            )
            if cig_input:
                st.session_state["cig_manuale"] = cig_input
                csa_data["cig"] = cig_input
                _salva_stato_cantiere()
        else:
            st.caption(f"CIG: {cig}")
    with col_cig2:
        if not cup or cup == "—":
            cup_input = st.text_input(
                "CUP (non rilevato — inserisci manualmente)",
                value=st.session_state.get("cup_manuale", ""),
                placeholder="Es: J24E22000060001",
                key="input_cup_manuale",
            )
            if cup_input:
                st.session_state["cup_manuale"] = cup_input
                csa_data["cup"] = cup_input
                _salva_stato_cantiere()
        else:
            st.caption(f"CUP: {cup}")

    st.divider()

    # Fila 1 — KPI principali
    durata = int(csa_data.get("durata_lavori_giorni") or 0)
    sal_tipo = csa_data.get("sal_tipo", "—")
    sal_gg = csa_data.get("sal_intervallo_giorni")

    dc = st.session_state.get("data_consegna_cantiere") or None
    avanz_pct = None
    giorni_trascorsi = None
    if dc and durata > 0:
        giorni_trascorsi = (date.today() - dc).days
        avanz_pct = min(100.0, max(0.0, giorni_trascorsi / durata * 100))

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        importo_str = f"€ {importo_netto:,.0f}" if importo_netto > 0 else csa_data.get("importo_lavori", "—")
        st.metric("💶 Importo netto", importo_str)
    with col2:
        st.metric("📅 Durata", f"{durata} gg" if durata else "—")
    with col3:
        sal_importo = csa_data.get("sal_importo_minimo_euro") or importo_netto
        sal_importo_val = _parse_importo(sal_importo) if sal_importo else importo_netto
        sal_detail = f"ogni {sal_gg} gg" if sal_gg else ""
        st.metric("📊 SAL", f"€ {sal_importo_val:,.0f}", delta=sal_detail, delta_color="off")
    with col4:
        if avanz_pct is not None:
            st.metric(
                "⏱️ Avanzamento temporale",
                f"{avanz_pct:.1f}%",
                delta=f"{giorni_trascorsi} gg trascorsi",
                delta_color="off",
            )
        else:
            st.metric("⏱️ Avanzamento temporale", "—", help="Imposta data consegna nel tab Calendario")

    # Fila 2 — KPI operativi
    oe = st.session_state.get("operatori_economici", {})
    subs = oe.get("subappaltatori", [])
    tot_sub = sum(float(s.get("importo", 0) or 0) for s in subs if s.get("tipo") == "subappalto")
    perc_sub = (tot_sub / importo_netto * 100) if importo_netto > 0 and tot_sub > 0 else None

    # Subaffidamenti (Art. 122 D.Lgs. 36/2023)
    tot_subaffid = sum(float(s.get("importo", 0) or 0) for s in subs if s.get("tipo") == "subaffidamento")
    perc_subaffid = (tot_subaffid / importo_netto * 100) if importo_netto > 0 and tot_subaffid > 0 else None

    registri = st.session_state.get("registri", {})
    sal_contab = registri.get("contabilita_sal", [])
    n_sal_pagati = len([s for s in sal_contab if s.get("stato_pagamento") == "pagato"])
    ordini_servizio = registri.get("ordini_servizio", [])
    n_ordini = len(ordini_servizio)

    if subs or sal_contab or ordini_servizio:
        st.divider()
        col_a, col_b, col_c, col_d = st.columns(4)
        with col_a:
            delta_sub = f"{perc_sub:.1f}% / 30% max" if perc_sub is not None else None
            st.metric("🏗️ Subappalti", f"€ {tot_sub:,.0f}", delta=delta_sub, delta_color="inverse" if (perc_sub or 0) > 30 else "off")
        with col_b:
            delta_subaffid = f"{perc_subaffid:.1f}% / 10% max" if perc_subaffid is not None else None
            st.metric("🏭 Subaffidamenti", f"€ {tot_subaffid:,.0f}", delta=delta_subaffid, delta_color="inverse" if (perc_subaffid or 0) > 10 else "off")
        with col_c:
            st.metric("📋 Ordini di Servizio", n_ordini)
        with col_d:
            st.metric("💰 SAL contabilità pagati", n_sal_pagati)

        if oe:
            render_durc_semaphore(oe)

    # Fila 3 — Subaffidamenti & Ordini di Servizio (da estrazione CSA)
    subaffid_importo = float(csa_data.get("subaffidamenti_importo", 0) or 0)
    subaffid_numero = int(csa_data.get("subaffidamenti_numero", 0) or 0)
    os_numero = int(csa_data.get("ordini_servizio_numero", 0) or 0)

    if subaffid_importo > 0 or subaffid_numero > 0 or os_numero > 0:
        st.divider()
        col_e, col_f, col_g, _ = st.columns(4)
        subaffid_limit = importo_netto * 0.10 if importo_netto > 0 else 0.0
        warn_sa = importo_netto > 0 and subaffid_importo > subaffid_limit > 0
        delta_sa = f"{subaffid_importo / importo_netto * 100:.1f}% (limite 10%)" if importo_netto > 0 else None
        with col_e:
            st.metric(
                "🔀 Subaffidamenti",
                f"€ {subaffid_importo:,.0f}",
                delta=delta_sa,
                delta_color="inverse" if warn_sa else "off",
                help="Art. 122 D.Lgs. 36/2023 — limite 10% importo subappalto",
            )
        with col_f:
            st.metric("👥 Subaffidatari", subaffid_numero)
        with col_g:
            st.metric("📝 Ordini di Servizio", os_numero, help="Ordini di servizio estratti dal CSA")
        if warn_sa:
            st.warning(
                f"⚠️ Subaffidamenti (€ {subaffid_importo:,.0f}) superano il limite del 10%"
                f" (max € {subaffid_limit:,.0f})"
            )

    # Avanzamento barra
    if avanz_pct is not None:
        st.divider()
        st.progress(avanz_pct / 100, text=f"Avanzamento temporale: {avanz_pct:.1f}%")

    # Alert riserve con scadenza imminente
    _riserve_all = st.session_state.get("registri", {}).get("riserve", [])
    _riserve_urgenti = []
    for _r in _riserve_all:
        if _r.get("stato", "").lower() == "iscritta" and _r.get("scadenza_esplicitazione"):
            try:
                _gg = (date.fromisoformat(_r["scadenza_esplicitazione"]) - date.today()).days
                if _gg <= 7:
                    _riserve_urgenti.append((_r.get("id", "—"), _r.get("lavorazione", _r.get("causale", "—")), _gg))
            except (ValueError, TypeError):
                pass
    if _riserve_urgenti:
        st.divider()
        st.markdown("#### ⚠️ Riserve — Scadenze imminenti")
        for _rid, _lav, _gg in _riserve_urgenti:
            _lav_short = str(_lav)[:50]
            if _gg <= 0:
                st.error(f"❌ Riserva **{_rid}** ({_lav_short}): termine esplicitazione **SCADUTO**")
            else:
                st.warning(f"⚠️ Riserva **{_rid}** ({_lav_short}): esplicitare entro **{_gg} giorni**")

    # Riepilogo rapido
    st.divider()
    col_r1, col_r2, col_r3 = st.columns(3)
    with col_r1:
        penale_perm = csa_data.get("penale_giornaliera_permille")
        if penale_perm and importo_netto > 0:
            penale_gg = importo_netto * float(penale_perm) / 1000
            st.info(f"⚠️ Penale giornaliera\n\n**€ {penale_gg:,.2f}** ({penale_perm}‰)")
        else:
            st.info(f"⚠️ Penale: {penale_perm or '—'}‰")
    with col_r2:
        collaudo = csa_data.get("collaudo_giorni")
        st.info(f"🏁 Collaudo entro\n\n**{collaudo} gg** dall'ultimazione" if collaudo else "🏁 Collaudo: —")
    with col_r3:
        tipo_c = csa_data.get("tipo_contratto", "—")
        prezzario = csa_data.get("prezzario_nome", "")
        anno = csa_data.get("prezzario_anno", "")
        st.info(f"📑 Tipo: **{tipo_c}**\n\n{prezzario} {anno}".strip())

    # Azioni rapide
    st.divider()
    st.markdown("#### Azioni rapide")
    ca1, ca2, ca3 = st.columns(3)
    with ca1:
        dc_input = st.date_input(
            "📅 Data consegna cantiere",
            value=st.session_state.get("data_consegna_cantiere") or date.today(),
            format="DD/MM/YYYY",
            key="dash_data_consegna",
        )
        if dc_input and dc_input != st.session_state.get("data_consegna_cantiere"):
            st.session_state.data_consegna_cantiere = dc_input
            aggiungi_log("Data consegna impostata", dc_input.isoformat(), tab="Dashboard")
            _salva_stato_cantiere()
            st.rerun()
    with ca2:
        ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
        st.metric("Ribasso d'asta", f"{ribasso_pct:.2f}%", help="Modifica dalla sidebar")
    with ca3:
        n_vp = len(st.session_state.get("_varianti_proroghe", []))
        if n_vp:
            giorni_vp = sum(int(v.get("giorni", 0) or 0) for v in st.session_state._varianti_proroghe)
            st.metric("Varianti/Proroghe", n_vp, delta=f"+{giorni_vp} gg" if giorni_vp > 0 else f"{giorni_vp} gg", delta_color="off")
        else:
            st.metric("Varianti/Proroghe", "Nessuna")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — SINTESI CSA
# ═══════════════════════════════════════════════════════════════════════════════

def _render_sintesi(csa_data: dict, importo_netto: float) -> None:
    st.header("📋 Sintesi CSA")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Dati generali")
        campi_gen = [
            ("Stazione Appaltante", csa_data.get("stazione_appaltante")),
            ("Tipo Lavori", csa_data.get("tipo_lavori")),
            ("Indirizzo", csa_data.get("indirizzo_cantiere")),
            ("Comune", csa_data.get("comune")),
            ("Provincia", csa_data.get("provincia")),
            ("Regione", csa_data.get("regione")),
        ]
        for label, val in campi_gen:
            if val:
                st.markdown(f"**{label}:** {val}")
        # CIG — editabile se non rilevato
        cig_s = csa_data.get("cig", "") or ""
        if cig_s and cig_s != "—":
            st.markdown(f"✅ **CIG:** {cig_s}")
        else:
            cig_s2 = st.text_input(
                "CIG (non rilevato — inserisci manualmente)",
                value=st.session_state.get("cig_manuale", ""),
                placeholder="Es: 9999999999",
                key="sintesi_cig_manuale",
            )
            if cig_s2:
                st.session_state["cig_manuale"] = cig_s2
                csa_data["cig"] = cig_s2
        # CUP — editabile se non rilevato
        cup_s = csa_data.get("cup", "") or ""
        if cup_s and cup_s != "—":
            st.markdown(f"✅ **CUP:** {cup_s}")
        else:
            cup_s2 = st.text_input(
                "CUP (non rilevato — inserisci manualmente)",
                value=st.session_state.get("cup_manuale", ""),
                placeholder="Es: J24E22000060001",
                key="sintesi_cup_manuale",
            )
            if cup_s2:
                st.session_state["cup_manuale"] = cup_s2
                csa_data["cup"] = cup_s2

    with col2:
        st.subheader("Parametri contrattuali")
        importo_lordo = _parse_importo(csa_data.get("importo_lavori"))
        ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
        st.markdown(f"**Importo lordo:** {csa_data.get('importo_lavori', '—')}")
        if ribasso_pct > 0:
            st.markdown(f"**Ribasso:** {ribasso_pct:.2f}%")
            st.markdown(f"**Importo netto:** € {importo_netto:,.2f}")
        st.markdown(f"**Durata:** {csa_data.get('durata_lavori_giorni', '—')} giorni")
        st.markdown(f"**Tipo contratto:** {csa_data.get('tipo_contratto', '—')}")
        prezzario = " ".join(filter(None, [csa_data.get("prezzario_nome"), csa_data.get("prezzario_anno")]))
        if prezzario:
            st.markdown(f"**Prezzario:** {prezzario}")
        st.markdown(f"**SAL tipo:** {csa_data.get('sal_tipo', '—')}")
        sal_gg = csa_data.get("sal_intervallo_giorni")
        sal_imp = csa_data.get("sal_importo_minimo_euro")
        sal_pct = csa_data.get("sal_percentuale_minima")
        if sal_gg:
            st.markdown(f"**SAL intervallo:** {sal_gg} giorni")
        if sal_imp:
            st.markdown(f"**SAL importo minimo:** € {sal_imp:,.2f}")
        if sal_pct:
            st.markdown(f"**SAL % minima:** {sal_pct}%")
        st.markdown(f"**Penale giornaliera:** {csa_data.get('penale_giornaliera_permille', '—')}‰")
        st.markdown(f"**Penale massima:** {csa_data.get('penale_massima_percentuale', '—')}%")
        st.markdown(f"**Riserve iscrizione:** {csa_data.get('riserve_iscrizione_giorni', '—')} gg")
        st.markdown(f"**Riserve quantificazione:** {csa_data.get('riserve_quantificazione_giorni', '—')} gg")
        st.markdown(f"**Collaudo:** {csa_data.get('collaudo_giorni', '—')} gg")
        oneri = csa_data.get("importo_oneri_sicurezza")
        if oneri:
            st.markdown(f"**Oneri sicurezza:** € {float(oneri):,.2f}")

    st.divider()

    # Categorie SOA
    soa_list = csa_data.get("categorie_soa", [])
    if soa_list:
        st.subheader("Categorie SOA")
        df_soa = pd.DataFrame(soa_list)
        col_map = {
            "codice": "Codice",
            "descrizione_categoria": "Descrizione",
            "classifica": "Classifica",
            "prevalente": "Prevalente",
            "motivazione": "Motivazione",
        }
        df_soa = df_soa.rename(columns=col_map)
        if "Prevalente" in df_soa.columns:
            df_soa["Prevalente"] = df_soa["Prevalente"].map(lambda x: "✅ Sì" if x else "No")
        st.dataframe(df_soa, use_container_width=True, hide_index=True)

    st.divider()

    # Subappalto
    st.subheader("Subappalto")
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        sub_pct = csa_data.get("subappalto_percentuale_massima")
        st.markdown(f"**Percentuale massima:** {sub_pct or 30}%")
        sub_aut = csa_data.get("subappalto_autorizzazione_richiesta")
        st.markdown(f"**Autorizzazione SA richiesta:** {'Sì' if sub_aut else 'No'}")
        sub_qual = csa_data.get("subappalto_qualificazione_richiesta")
        st.markdown(f"**Qualificazione SOA richiesta:** {'Sì' if sub_qual else 'No'}")
    with col_s2:
        vietate = csa_data.get("subappalto_categorie_vietate", [])
        if vietate:
            st.markdown(f"**Categorie vietate:** {', '.join(vietate)}")
        note_sub = csa_data.get("subappalto_note")
        if note_sub:
            st.markdown(f"**Note:** {note_sub}")

    st.divider()

    # Obblighi
    col_ob1, col_ob2 = st.columns(2)
    with col_ob1:
        st.subheader("Obblighi Appaltatore")
        for ob in csa_data.get("obblighi_appaltatore", []):
            st.markdown(f"• {ob}")
    with col_ob2:
        st.subheader("Obblighi Stazione Appaltante")
        for ob in csa_data.get("obblighi_stazione_appaltante", []):
            st.markdown(f"• {ob}")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CHECKLIST
# ═══════════════════════════════════════════════════════════════════════════════

def _render_checklist(csa_data: dict, api_key: str) -> None:
    st.header("✅ Checklist Attività")

    if "checklist_stato" not in st.session_state:
        st.session_state.checklist_stato = {}

    dc = st.session_state.get("data_consegna_cantiere") or None

    # Filtro stato
    filtro = st.selectbox(
        "Mostra", ["Tutti"] + _STATI_CHECKLIST, key="ck_filtro_stato"
    )

    for cat_label, cat_key in _CHECKLIST_CATEGORIE.items():
        voci = csa_data.get(cat_key, [])
        if not voci:
            continue

        with st.expander(f"📁 {cat_label} ({len(voci)} voci)", expanded=(cat_label == "Prime settimane")):
            stato_cat = st.session_state.checklist_stato.setdefault(cat_key, {})

            for idx, voce in enumerate(voci):
                attivita = voce.get("attivita", "")
                priorita = voce.get("priorita", "media")
                termine_gg = voce.get("termine_giorni")

                stato_corrente = (stato_cat.get(str(idx)) or {}).get("stato", "Da fare")

                # Calcola se in ritardo
                if termine_gg and dc and stato_corrente not in ("Completato",):
                    scadenza = dc + timedelta(days=termine_gg)
                    if scadenza < date.today() and stato_corrente == "Da fare":
                        stato_corrente = "In ritardo"
                        stato_cat.setdefault(str(idx), {})["stato"] = stato_corrente

                if filtro != "Tutti" and stato_corrente != filtro:
                    continue

                icona_prior = {"alta": "🔴", "media": "🟡", "bassa": "🟢"}.get(priorita, "⚪")
                icona_stato = _STATI_COLORI.get(stato_corrente, "⚪")

                col_st, col_att, col_ref, col_doc = st.columns([2, 4, 3, 2])
                with col_st:
                    nuovo_stato = st.selectbox(
                        "Stato",
                        _STATI_CHECKLIST,
                        index=_STATI_CHECKLIST.index(stato_corrente) if stato_corrente in _STATI_CHECKLIST else 0,
                        key=f"ck_{cat_key}_{idx}_stato",
                        label_visibility="collapsed",
                    )
                    if nuovo_stato != stato_corrente:
                        stato_cat.setdefault(str(idx), {})["stato"] = nuovo_stato
                        aggiungi_log("Checklist aggiornata", f"{attivita[:50]} → {nuovo_stato}", tab="Checklist")
                        _salva_stato_cantiere()
                        st.rerun()
                with col_att:
                    term_str = f" (entro {termine_gg}gg)" if termine_gg else ""
                    st.markdown(f"{icona_prior} **{attivita}**{term_str}")
                    if termine_gg and dc:
                        scadenza_data = dc + timedelta(days=termine_gg)
                        st.caption(f"Scadenza: {scadenza_data.strftime('%d/%m/%Y')}")
                with col_ref:
                    voce_stato = stato_cat.setdefault(str(idx), {})
                    ref_nome = st.text_input(
                        "Referente",
                        value=voce_stato.get("referente_nome", ""),
                        key=f"ck_{cat_key}_{idx}_ref",
                        placeholder="Nome referente",
                        label_visibility="collapsed",
                    )
                    if ref_nome != voce_stato.get("referente_nome", ""):
                        voce_stato["referente_nome"] = ref_nome
                        _salva_stato_cantiere()
                with col_doc:
                    doc_nome = voce_stato.get("documento_nome", "")
                    if doc_nome:
                        doc_path = RESULTS_DIR / "documenti" / doc_nome
                        render_doc_buttons(doc_path, key=f"ck_{cat_key}_{idx}_docbtn")
                    up = st.file_uploader(
                        "Allega",
                        type=_TIPI_FILE_DOC,
                        key=f"ck_{cat_key}_{idx}_up",
                        label_visibility="collapsed",
                    )
                    if up is not None and up.name != doc_nome:
                        dest_dir = RESULTS_DIR / "documenti"
                        dest_dir.mkdir(parents=True, exist_ok=True)
                        (dest_dir / up.name).write_bytes(up.getvalue())
                        voce_stato["documento_nome"] = up.name
                        aggiungi_log("Documento allegato checklist", up.name, tab="Checklist")
                        _salva_stato_cantiere()
                        st.rerun()

                st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — DOCUMENTI
# ═══════════════════════════════════════════════════════════════════════════════

def _render_documenti(csa_data: dict, api_key: str) -> None:
    st.header("📄 Documenti di Progetto")

    # Migrazione formato vecchio → nuovo
    if "doc_elaborati" not in st.session_state:
        st.session_state.doc_elaborati = {"grafico": [], "tecnico": [], "amministrativo": []}

    doc_elaborati = st.session_state.doc_elaborati

    # Popola con elaborati da CSA se ancora vuoto
    elaborati_csa = csa_data.get("elaborati", [])
    if elaborati_csa and not any(doc_elaborati.values()):
        for el in elaborati_csa:
            cat = el.get("categoria", "tecnico")
            if cat not in doc_elaborati:
                cat = "tecnico"
            if not any(d.get("codice") == el.get("codice") for d in doc_elaborati[cat]):
                doc_elaborati[cat].append({
                    "codice": el.get("codice", ""),
                    "titolo": el.get("titolo", ""),
                    "is_variante": False,
                    "data_approvazione_variante": "",
                    "note_variante": "",
                    "path": "",
                    "data_upload": "",
                })

    # Repository elaborati — 3 sezioni
    categorie_info = {
        "grafico": ("🗺️ Elaborati Grafici", "Tavole, planimetrie, sezioni, particolari costruttivi"),
        "tecnico": ("📐 Elaborati Tecnici", "Relazioni, computo metrico, elenco prezzi, PSC, cronoprogramma"),
        "amministrativo": ("📁 Elaborati Amministrativi", "Schema contratto, capitolato, disciplinare"),
    }

    RESULTS_ELAB = RESULTS_DIR / "elaborati"

    for cat_key, (cat_label, cat_desc) in categorie_info.items():
        with st.expander(f"{cat_label}", expanded=True):
            st.caption(cat_desc)

            docs_cat = doc_elaborati.get(cat_key, [])
            if not docs_cat:
                st.info("Nessun elaborato in questa categoria.")
            else:
                for i, doc in enumerate(docs_cat):
                    col_cod, col_tit, col_var, col_stato, col_act = st.columns([1.5, 4, 1.5, 1, 2])
                    with col_cod:
                        st.markdown(f"`{doc.get('codice', '—')}`")
                    with col_tit:
                        st.markdown(f"**{doc.get('titolo', '—')}**")
                    with col_var:
                        if doc.get("is_variante"):
                            st.markdown("🔴 **Variante**")
                    with col_stato:
                        if doc.get("path"):
                            st.success("✅")
                        else:
                            st.warning("—")
                    with col_act:
                        if doc.get("path"):
                            render_doc_buttons(pathlib.Path(doc["path"]), key=f"doc_{cat_key}_{i}_btn")
                        up = st.file_uploader(
                            "Carica",
                            type=["pdf"],
                            key=f"doc_{cat_key}_{i}_up",
                            label_visibility="collapsed",
                        )
                        if up is not None:
                            path_corrente = doc.get("path", "")
                            gia_caricato = bool(path_corrente) and str(path_corrente).endswith(up.name)
                            if not gia_caricato:
                                RESULTS_ELAB.mkdir(parents=True, exist_ok=True)
                                dest = RESULTS_ELAB / f"{cat_key}_{doc.get('codice', str(i))}_{up.name}"
                                dest.write_bytes(up.getvalue())
                                doc_elaborati[cat_key][i]["path"] = str(dest)
                                doc_elaborati[cat_key][i]["data_upload"] = datetime.now().isoformat()
                                aggiungi_log("Elaborato caricato", f"{doc.get('codice')} {up.name}", tab="Documenti")
                                _salva_stato_cantiere()
                                st.rerun()

                    # Ricerca testo nel PDF
                    if doc.get("path"):
                        cerca_key = f"doc_{cat_key}_{i}_cerca"
                        query = st.text_input("🔍 Cerca nel PDF", key=cerca_key, placeholder="Testo da cercare…", label_visibility="collapsed")
                        if query.strip():
                            try:
                                import fitz
                                pdf_doc = fitz.open(doc["path"])
                                trovati = []
                                for n_pag, pag in enumerate(pdf_doc):
                                    if query.lower() in pag.get_text().lower():
                                        trovati.append(n_pag + 1)
                                pdf_doc.close()
                                if trovati:
                                    st.caption(f"Trovato alle pagine: {trovati}")
                                else:
                                    st.caption("Testo non trovato.")
                            except Exception:
                                st.caption("Ricerca non disponibile.")
                    st.markdown("---")

            # Form aggiungi documento
            with st.expander("➕ Aggiungi documento"):
                c_cod, c_tit, c_var = st.columns(3)
                with c_cod:
                    new_cod = st.text_input("Codice", key=f"doc_{cat_key}_new_cod", placeholder="TAV.01")
                with c_tit:
                    new_tit = st.text_input("Titolo", key=f"doc_{cat_key}_new_tit")
                with c_var:
                    new_var = st.checkbox("È variante", key=f"doc_{cat_key}_new_var")
                if st.button("Aggiungi", key=f"doc_{cat_key}_new_btn"):
                    if new_tit.strip():
                        doc_elaborati[cat_key].append({
                            "codice": new_cod.strip(),
                            "titolo": new_tit.strip(),
                            "is_variante": new_var,
                            "data_approvazione_variante": "",
                            "note_variante": "",
                            "path": "",
                            "data_upload": "",
                        })
                        _salva_stato_cantiere()
                        st.rerun()

    st.divider()

    # Redazione documenti ufficiali
    with st.expander("✍️ Redazione Documenti Ufficiali", expanded=False):
        if not api_key:
            st.warning("API Key richiesta per generare documenti.")
            return

        tipo_doc = st.selectbox(
            "Tipo documento",
            ["riserva", "verbale_consegna", "proroga", "contestazione"],
            format_func=lambda x: {
                "riserva": "📋 Riserva art.120",
                "verbale_consegna": "📝 Verbale Consegna",
                "proroga": "⏳ Richiesta Proroga",
                "contestazione": "⚠️ Nota Contestazione",
            }[x],
            key="doc_tipo_redazione",
        )

        data_lettera = st.date_input("Data documento", value=date.today(), key="doc_data_lettera", format="DD/MM/YYYY")

        params = {"data_lettera": data_lettera.strftime("%d/%m/%Y")}

        if tipo_doc == "riserva":
            params["motivo"] = st.text_area("Motivo della riserva", key="doc_riserva_motivo", height=80)
            params["importo"] = st.text_input("Importo indicativo", key="doc_riserva_importo", placeholder="€ 25.000,00")
            params["data_evento"] = st.text_input("Data evento generatore", key="doc_riserva_data_evento")
        elif tipo_doc == "verbale_consegna":
            params["osservazioni"] = st.text_area("Osservazioni appaltatore", key="doc_vc_osservazioni", height=80)
        elif tipo_doc == "proroga":
            params["causa"] = st.text_area("Causa della proroga", key="doc_proroga_causa", height=80)
            params["giorni"] = st.text_input("Giorni richiesti", key="doc_proroga_giorni")
            params["documentazione"] = st.text_input("Documentazione allegata", key="doc_proroga_docs")
        elif tipo_doc == "contestazione":
            params["oggetto"] = st.text_input("Oggetto contestazione", key="doc_cont_oggetto")
            params["fatti"] = st.text_area("Descrizione fatti", key="doc_cont_fatti", height=80)
            params["richiesta"] = st.text_area("Richiesta formale", key="doc_cont_richiesta", height=60)

        lunghezza = st.select_slider(
            "Lunghezza documento",
            options=["Breve", "Standard", "Dettagliato"],
            value="Standard",
            key="doc_lunghezza",
        )
        max_tok = {"Breve": 512, "Standard": 1500, "Dettagliato": 2500}[lunghezza]

        key_doc = f"_doc_{tipo_doc}"
        if st.button("✨ Genera documento", type="primary", key=f"doc_genera_{tipo_doc}"):
            with st.spinner("Claude Haiku sta redigendo il documento…"):
                try:
                    testo = generate_document(tipo_doc, csa_data, params, api_key, max_tokens=max_tok)
                    st.session_state[key_doc] = testo
                    aggiungi_log("Documento generato", tipo_doc, tab="Documenti")
                except Exception as e:
                    st.error(f"Errore generazione: {e}")

        if st.session_state.get(key_doc):
            st.text_area("Documento generato", value=st.session_state[key_doc], height=300, key=f"doc_output_{tipo_doc}")
            st.download_button(
                "⬇️ Scarica .txt",
                data=st.session_state[key_doc].encode("utf-8"),
                file_name=f"{tipo_doc}_{date.today().isoformat()}.txt",
                mime="text/plain",
                key=f"doc_dl_{tipo_doc}",
            )


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — CALENDARIO
# ═══════════════════════════════════════════════════════════════════════════════

def _render_calendario(csa_data: dict, details: dict, importo_netto: float) -> None:
    st.header("📅 Calendario Contrattuale")

    # Data consegna condivisa
    dc_corrente = st.session_state.get("data_consegna_cantiere") or date.today()
    nuova_dc = st.date_input(
        "📅 Data consegna cantiere",
        value=dc_corrente,
        format="DD/MM/YYYY",
        key="cal_data_consegna",
    )
    if nuova_dc and nuova_dc != st.session_state.get("data_consegna_cantiere"):
        st.session_state.data_consegna_cantiere = nuova_dc
        aggiungi_log("Data consegna aggiornata", nuova_dc.isoformat(), tab="Calendario")
        _salva_stato_cantiere()
        st.rerun()

    dc = nuova_dc or date.today()

    # Parametri SAL
    durata_base = int(details.get("durata_lavori_giorni") or 180)
    sal_tipo = details.get("sal_tipo", "tempo")
    sal_gg = details.get("sal_intervallo_giorni")
    sal_imp_base = details.get("sal_importo_minimo_euro")
    sal_pct = details.get("sal_percentuale_minima")

    sal_imp = sal_imp_base
    if not sal_imp and sal_pct and importo_netto > 0:
        sal_imp = importo_netto * sal_pct / 100

    # Varianti e Proroghe
    st.divider()
    st.subheader("🔄 Varianti e Proroghe")

    if "_varianti_proroghe" not in st.session_state:
        st.session_state._varianti_proroghe = []

    vp_list = st.session_state._varianti_proroghe
    giorni_varianti = sum(int(v.get("giorni", 0) or 0) for v in vp_list)
    durata_eff = durata_base + giorni_varianti

    if vp_list:
        for vp_idx, vp in enumerate(vp_list):
            col_t, col_d, col_gg, col_del = st.columns([2, 2, 1, 1])
            with col_t:
                st.markdown(f"**{vp.get('tipo', '')}**")
                st.caption(vp.get("descrizione", ""))
            with col_d:
                st.caption(f"Approvazione: {vp.get('data_approvazione', '—')}")
            with col_gg:
                gg_vp = int(vp.get("giorni", 0) or 0)
                delta_str = f"+{gg_vp} gg" if gg_vp >= 0 else f"{gg_vp} gg"
                st.metric("Giorni", delta_str)
            with col_del:
                if st.button("🗑️", key=f"vp_del_{vp_idx}"):
                    vp_list.pop(vp_idx)
                    _salva_stato_cantiere()
                    st.rerun()
        st.info(f"Durata contrattuale: {durata_base} gg + {giorni_varianti} gg varianti/proroghe = **{durata_eff} gg totali**")
        nuova_fine = dc + timedelta(days=durata_eff)
        st.success(f"Nuova scadenza: **{nuova_fine.strftime('%d/%m/%Y')}**")
    else:
        st.caption("Nessuna variante o proroga registrata.")

    with st.expander("➕ Aggiungi variante / proroga"):
        col_vp1, col_vp2, col_vp3 = st.columns(3)
        with col_vp1:
            tipo_vp = st.selectbox("Tipo", ["Variante", "Proroga", "Sospensione tecnica"], key="cal_vp_tipo")
            data_appr_vp = st.date_input("Data approvazione", value=date.today(), key="cal_vp_data", format="DD/MM/YYYY")
        with col_vp2:
            giorni_vp = st.number_input("Giorni (±)", value=0, step=1, key="cal_vp_giorni", help="Positivo per proroga, negativo per anticipo")
            descr_vp = st.text_input("Descrizione", key="cal_vp_descr")
        with col_vp3:
            # Preview
            if giorni_vp != 0 or descr_vp:
                nuova_fine_prev = dc + timedelta(days=durata_eff + int(giorni_vp))
                st.info(f"Nuova scadenza prevista:\n**{nuova_fine_prev.strftime('%d/%m/%Y')}**")
        if st.button("Aggiungi", key="cal_vp_add_btn"):
            vp_list.append({
                "tipo": tipo_vp,
                "descrizione": descr_vp,
                "data_approvazione": data_appr_vp.isoformat(),
                "giorni": int(giorni_vp),
            })
            aggiungi_log(f"{tipo_vp} aggiunta", f"{descr_vp} ({'+' if giorni_vp >= 0 else ''}{giorni_vp} gg)", tab="Calendario")
            _salva_stato_cantiere()
            st.rerun()

    # Sospensioni
    st.divider()
    st.subheader("⏸️ Sospensioni")

    sospensioni: list[Sospensione] = st.session_state.get("sospensioni", [])

    with st.expander("➕ Aggiungi sospensione"):
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            data_inizio_s = st.date_input("Data inizio", value=date.today(), key="cal_sosp_inizio", format="DD/MM/YYYY")
            tipo_s = st.selectbox("Tipo", ["totale", "parziale"], key="cal_sosp_tipo")
        with col_s2:
            data_fine_s = st.date_input("Data fine (opzionale)", value=None, key="cal_sosp_fine", format="DD/MM/YYYY")
            perc_s = st.slider("Percentuale (se parziale)", 1, 99, 50, key="cal_sosp_perc", disabled=(tipo_s == "totale"))
        motivo_s = st.text_input("Motivo", key="cal_sosp_motivo")
        if st.button("Aggiungi sospensione", key="cal_sosp_add"):
            if not motivo_s.strip():
                st.error("Il motivo è obbligatorio.")
            else:
                st.session_state._sosp_counter = st.session_state.get("_sosp_counter", 0) + 1
                sospensioni.append(Sospensione(
                    id=st.session_state._sosp_counter,
                    data_inizio=data_inizio_s,
                    data_fine=data_fine_s,
                    tipo=tipo_s,
                    motivo=motivo_s.strip(),
                    percentuale=float(perc_s) if tipo_s == "parziale" else 100.0,
                ))
                st.session_state.sospensioni = sospensioni
                aggiungi_log("Sospensione aggiunta", f"{tipo_s}: {motivo_s}", tab="Calendario")
                _salva_stato_cantiere()
                st.rerun()

    if sospensioni:
        for sosp in sospensioni:
            col_sv1, col_sv2, col_sv3 = st.columns([3, 2, 1])
            with col_sv1:
                st.markdown(f"**{sosp.tipo.upper()}** — {sosp.motivo}")
            with col_sv2:
                fine_str = sosp.data_fine.strftime("%d/%m/%Y") if sosp.data_fine else "In corso"
                st.caption(f"{sosp.data_inizio.strftime('%d/%m/%Y')} → {fine_str}")
            with col_sv3:
                if st.button("🗑️", key=f"cal_sosp_del_{sosp.id}"):
                    st.session_state.sospensioni = [s for s in sospensioni if s.id != sosp.id]
                    _salva_stato_cantiere()
                    st.rerun()

    # Calcolo e visualizzazione calendario
    st.divider()
    st.subheader("📋 Calendario eventi")

    cal = calcola_calendario(
        data_consegna=dc,
        durata_giorni=durata_eff,
        importo_contratto=importo_netto,
        sal_tipo=sal_tipo,
        sal_intervallo_giorni=sal_gg,
        sal_importo_soglia=sal_imp,
        sospensioni=sospensioni,
    )

    col_cal1, col_cal2, col_cal3 = st.columns(3)
    col_cal1.metric("Data fine lavori", cal["data_fine_lavori"].strftime("%d/%m/%Y"))
    col_cal2.metric("Durata effettiva", f"{cal['durata_effettiva_giorni']} gg")
    col_cal3.metric("Giorni sospesi", cal["giorni_sospesi_totali"])

    oggi = date.today()
    eventi_df = []
    for ev in cal["eventi"]:
        delta = (ev["data"] - oggi).days
        if delta < 0:
            stato = "✅ Passato"
        elif delta == 0:
            stato = "⚡ Oggi"
        elif delta <= 30 and ev.get("critico"):
            stato = "🔴 Critico"
        else:
            stato = "🔵 Futuro"
        eventi_df.append({
            "Data": ev["data"].strftime("%d/%m/%Y"),
            "Tipo": ev["tipo"],
            "Descrizione": ev["descrizione"],
            "Giorno": ev["giorno_contratto"],
            "Stato": stato,
        })

    df_cal = pd.DataFrame(eventi_df)
    st.dataframe(df_cal, use_container_width=True, hide_index=True)

    # Export PDF calendario
    if st.button("📄 Esporta PDF calendario", key="cal_pdf_export"):
        try:
            pdf_bytes_cal = _genera_pdf_calendario(cal, csa_data)
            st.download_button(
                "⬇️ Scarica PDF",
                data=pdf_bytes_cal,
                file_name=f"calendario_{date.today().isoformat()}.pdf",
                mime="application/pdf",
                key="cal_pdf_dl",
            )
        except ImportError:
            st.error("fpdf2 non disponibile: pip install fpdf2")
        except Exception as e:
            st.error(f"Errore PDF: {e}")


def _genera_pdf_calendario(cal: dict, csa_data: dict) -> bytes:
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    _fn = "DejaVu"
    if os.path.isfile(_FONT_REGULAR):
        pdf.add_font("DejaVu", "", _FONT_REGULAR)
    if os.path.isfile(_FONT_BOLD):
        pdf.add_font("DejaVu", "B", _FONT_BOLD)
    if not os.path.isfile(_FONT_REGULAR):
        _fn = "Helvetica"

    pdf.set_font(_fn, "B", 13)
    pdf.cell(0, 8, "Calendario Contrattuale — DTC", ln=True, align="C")
    pdf.set_font(_fn, "", 9)
    pdf.cell(0, 5, f"{csa_data.get('tipo_lavori', '')} — {csa_data.get('comune', '')} ({csa_data.get('provincia', '')})", ln=True, align="C")
    pdf.ln(3)

    hdrs = [("Data", 28), ("Tipo", 40), ("Descrizione", 155), ("Giorno", 18)]
    pdf.set_fill_color(31, 78, 121)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(_fn, "B", 8)
    for h, w in hdrs:
        pdf.cell(w, 7, h, border=1, fill=True, align="C")
    pdf.ln()

    oggi = date.today()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font(_fn, "", 7)
    for ev in cal["eventi"]:
        delta = (ev["data"] - oggi).days
        if delta < 0:
            pdf.set_fill_color(230, 230, 230)
        elif delta <= 30 and ev.get("critico"):
            pdf.set_fill_color(255, 220, 220)
        else:
            pdf.set_fill_color(220, 240, 255)
        row_data = [
            (ev["data"].strftime("%d/%m/%Y"), 28),
            (ev["tipo"], 40),
            (ev["descrizione"][:80], 155),
            (str(ev["giorno_contratto"]), 18),
        ]
        for val, w in row_data:
            pdf.cell(w, 6, val, border=1, fill=True, align="L")
        pdf.ln()

    return bytes(pdf.output())


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 5 — PENALI & REVISIONE PREZZI
# ═══════════════════════════════════════════════════════════════════════════════

def _render_penali(csa_data: dict, details: dict, importo_netto: float) -> None:
    st.header("💰 Penali & Revisione Prezzi")

    penale_perm = float(details.get("penale_giornaliera_permille") or 1.0)
    penale_max_pct = float(details.get("penale_massima_percentuale") or 10.0)
    oneri_sic = float(details.get("importo_oneri_sicurezza") or 0.0)

    col_p1, col_p2 = st.columns(2)

    # Calcolo penali
    with col_p1:
        st.subheader("⚠️ Calcolo Penali")
        importo_pen = st.number_input(
            "Importo contrattuale (€)",
            min_value=0.0,
            value=importo_netto if importo_netto > 0 else 0.0,
            step=1000.0,
            format="%.2f",
            key="pen_importo",
        )
        permille = st.number_input(
            "Penale giornaliera (‰)",
            min_value=0.0,
            value=penale_perm,
            step=0.1,
            format="%.2f",
            key="pen_permille",
        )
        giorni_ritardo = st.number_input(
            "Giorni di ritardo",
            min_value=0,
            value=0,
            step=1,
            key="pen_giorni",
        )
        max_pct = st.number_input(
            "Penale massima (%)",
            min_value=0.0,
            value=penale_max_pct,
            step=0.5,
            format="%.2f",
            key="pen_max_pct",
        )

        if importo_pen > 0:
            res_pen = calcola_penale_cumulativa(
                importo_contratto=importo_pen,
                permille=permille,
                giorni_ritardo=int(giorni_ritardo),
                penale_massima_percent=max_pct,
            )
            st.metric("Penale giornaliera", f"€ {res_pen['penale_giornaliera_euro']:,.2f}")
            st.metric(
                "Penale cumulativa",
                f"€ {res_pen['penale_cumulativa_euro']:,.2f}",
                delta="CAP RAGGIUNTO" if res_pen["cap_raggiunto"] else None,
                delta_color="inverse" if res_pen["cap_raggiunto"] else "off",
            )
            st.metric("Penale massima", f"€ {res_pen['penale_massima_euro']:,.2f}")
            if res_pen["giorni_al_cap"]:
                st.caption(f"Il cap si raggiunge al giorno {res_pen['giorni_al_cap']}")
            aggiungi_log(
                "Penale calcolata",
                f"{giorni_ritardo} gg — € {res_pen['penale_cumulativa_euro']:,.2f}",
                tab="Penali",
            )

    # Revisione prezzi
    with col_p2:
        st.subheader("📈 Revisione Prezzi (art.60)")
        importo_rev = st.number_input(
            "Importo contrattuale (€)",
            min_value=0.0,
            value=importo_netto if importo_netto > 0 else 0.0,
            step=1000.0,
            format="%.2f",
            key="rev_importo",
        )
        oneri_rev = st.number_input(
            "Oneri sicurezza (€)",
            min_value=0.0,
            value=oneri_sic,
            step=100.0,
            format="%.2f",
            key="rev_oneri",
        )
        idx_offerta = st.number_input(
            "Indice ISTAT all'offerta",
            min_value=0.01,
            value=100.0,
            step=0.1,
            format="%.2f",
            key="rev_idx_offerta",
        )
        idx_aggiorn = st.number_input(
            "Indice ISTAT aggiornato",
            min_value=0.01,
            value=105.0,
            step=0.1,
            format="%.2f",
            key="rev_idx_aggiorn",
        )
        franchigia = st.number_input(
            "Franchigia (%)",
            min_value=0.0,
            value=5.0,
            step=0.5,
            format="%.2f",
            key="rev_franchigia",
        )

        if importo_rev > 0:
            res_rev = simula_revisione_prezzi(
                importo_contratto=importo_rev,
                oneri_sicurezza=oneri_rev,
                indice_istat_offerta=idx_offerta,
                indice_istat_aggiornamento=idx_aggiorn,
                soglia_franchigia_percent=franchigia,
            )
            if "error" in res_rev:
                st.error(res_rev["error"])
            else:
                st.metric("Variazione indice", f"{res_rev['variazione_totale_percent']:+.2f}%")
                st.metric(
                    "Compensazione",
                    f"€ {abs(res_rev['compensazione_euro']):,.2f}",
                    delta="a favore appaltatore" if res_rev["a_favore_appaltatore"] else "a favore SA",
                    delta_color="normal" if res_rev["a_favore_appaltatore"] else "inverse",
                )
                st.caption(res_rev["note"])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 — MAPPA FORNITORI
# ═══════════════════════════════════════════════════════════════════════════════

def _render_mappa(csa_data: dict) -> None:
    from modules.geocoder import geocode_site
    from modules.supplier_search import search_suppliers, CATEGORY_QUERIES, CATEGORY_KEYWORDS
    from modules.map_renderer import build_map

    st.header("🗺️ Mappa Fornitori")

    # Geocoding
    coords = st.session_state.get("coords")
    if coords is None and not st.session_state.get("_demo_active"):
        if st.button("📍 Geolocalizza cantiere", key="mappa_geocode"):
            with st.spinner("Nominatim geocoding…"):
                coords = geocode_site(csa_data)
                if coords:
                    st.session_state.coords = coords
                    aggiungi_log("Cantiere geolocalizzato", f"({coords[0]:.4f}, {coords[1]:.4f})", tab="Mappa")
                    st.rerun()
                else:
                    st.error("Geocoding non riuscito. Inserisci coordinate manualmente.")

    if coords is None:
        col_lat, col_lon = st.columns(2)
        with col_lat:
            lat_man = st.number_input("Latitudine", value=45.46, step=0.001, format="%.4f", key="mappa_lat_man")
        with col_lon:
            lon_man = st.number_input("Longitudine", value=9.19, step=0.001, format="%.4f", key="mappa_lon_man")
        if st.button("Usa coordinate", key="mappa_use_man"):
            st.session_state.coords = (lat_man, lon_man)
            st.rerun()
        return

    lat, lon = coords
    st.caption(f"Cantiere: {lat:.4f}N, {lon:.4f}E")

    # Ricerca fornitori
    col_m1, col_m2 = st.columns([2, 1])
    with col_m1:
        raggio = st.slider("Raggio di ricerca (km)", 10, 100, 50, key="mappa_raggio")
    with col_m2:
        categorie_sel = st.multiselect(
            "Categorie",
            options=list(CATEGORY_QUERIES.keys()),
            default=list(CATEGORY_QUERIES.keys())[:5],
            key="mappa_categorie",
        )

    suppliers = st.session_state.get("suppliers")
    if not st.session_state.get("_demo_active"):
        if st.button("🔍 Cerca fornitori", type="primary", key="mappa_cerca"):
            with st.spinner(f"Overpass API — {len(categorie_sel)} categorie nel raggio di {raggio} km…"):
                suppliers = search_suppliers(lat, lon, raggio, categorie_sel)
                st.session_state.suppliers = suppliers
                aggiungi_log("Ricerca fornitori", f"{len(suppliers)} trovati nel raggio {raggio}km", tab="Mappa")
                st.rerun()

    if suppliers is not None:
        n_obt = len(suppliers)
        st.info(f"**{n_obt}** fornitori trovati nell'area")

        if not suppliers.empty:
            m = build_map(lat, lon, suppliers, raggio)
            try:
                from streamlit_folium import st_folium
                st_folium(m, height=500, use_container_width=True)
            except ImportError:
                st.warning("Installa streamlit-folium: pip install streamlit-folium")

            # Tabella fornitori
            cols_show = ["Categoria", "Nome", "Indirizzo", "Telefono", "Email", "Sito Web"]
            cols_available = [c for c in cols_show if c in suppliers.columns]
            st.dataframe(suppliers[cols_available], use_container_width=True, hide_index=True)

        # Link Google Maps / Pagine Gialle per ogni categoria
        st.divider()
        st.markdown("#### Link ricerca per categoria")
        comune = csa_data.get("comune", "")
        for cat, keyword in CATEGORY_KEYWORDS.items():
            query_gm = f"{keyword} {comune} Italy".replace(" ", "+")
            query_pg = keyword.replace(" ", "+")
            prov = csa_data.get("provincia", "")
            st.markdown(
                f"**{cat}** — "
                f"[Google Maps](https://www.google.com/maps/search/{query_gm}) | "
                f"[Pagine Gialle](https://www.paginegialle.it/ricerca/{query_pg}/{comune.lower()}-{prov.lower()})"
            )


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 7 — GUIDA
# ═══════════════════════════════════════════════════════════════════════════════

def _render_guida() -> None:
    st.header("❓ Guida all'utilizzo")

    st.markdown("""
## Strumento DTC — Direttore Tecnico di Cantiere

Strumento web per il DTC di appalti pubblici italiani regolati da **D.Lgs. 36/2023** e **DM 49/2018**.

---

### 🚀 Come iniziare

1. **Inserisci la Claude API Key** nella sidebar (sk-ant-…)
2. **Carica il PDF del CSA** (Capitolato Speciale d'Appalto)
3. Clicca **"Avvia analisi CSA"** — Smart Extract filtra le pagine rilevanti
4. Esplora i **11 tab** con tutte le funzionalità

**Senza API Key:** usa la **modalità Demo** (dati fittizi cantiere Bergamo BG)

---

### 📑 Tab disponibili

| Tab | Funzione |
|-----|----------|
| 🏠 Dashboard | KPI principali, avanzamento temporale, azioni rapide |
| 📋 Sintesi CSA | Dati estratti: categorie SOA, parametri contrattuali, obblighi |
| ✅ Checklist | Attività prime settimane, sicurezza, assicurative, materiali |
| 📄 Documenti | Repository elaborati, upload PDF per codice, ricerca testo, redazione documenti |
| 📅 Calendario | SAL (tempo/importo/misto), sospensioni, varianti/proroghe |
| 💰 Penali | Calcolo penali per ritardo + revisione prezzi art.60 |
| 🏢 Operatori | Appaltatore L0, subappaltatori L1, subaffidatari L2, semaforo DURC |
| 🗂️ Registri | Riserve, Verbali, Non Conformità, Ordini di Servizio, Contabilità SAL |
| 🗺️ Mappa | Geocoding Nominatim + ricerca Overpass API + mappa Folium |
| 📊 Pianificazione | Cronoprogramma Gantt + CME + associazione fasi → voci + export Excel |
| 📋 Log | Log cronologico operazioni con filtri ed export PDF |

---

### 💡 Funzionalità globali (sidebar)

- **Ribasso d'asta** — si propaga a Calendario, Penali, Revisione prezzi, Export
- **Export Excel** — 4 fogli: Sintesi, Scadenze/Obblighi, Checklist, Calendario
- **Notifiche email** — SMTP/STARTTLS con scadenze critiche entro N giorni
- **Analisi salvate** — JSON in `results/`, carica senza rifare l'analisi API

---

### 📋 Note legali

Questo strumento è un **supporto operativo** per il DTC. Non sostituisce la consulenza legale professionale. Tutti i calcoli e i documenti generati devono essere verificati da un professionista qualificato prima dell'uso ufficiale.

Le interpretazioni normative (D.Lgs. 36/2023, DM 49/2018) si basano sul testo vigente alla data di sviluppo. Verificare sempre l'aggiornamento normativo.
""")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    _init_session_state()
    api_key = _render_sidebar()

    csa_data = st.session_state.get("csa_data")

    if csa_data is None:
        # Schermata di benvenuto
        st.markdown("# 🏗️ Strumento DTC")
        st.markdown("**Direttore Tecnico di Cantiere — Appalti Pubblici Italiani**")
        st.markdown("D.Lgs. 36/2023 · DM 49/2018 · Allegato II.12")
        st.divider()

        col_w1, col_w2, col_w3 = st.columns(3)
        with col_w1:
            st.info("**1️⃣ API Key**\n\nInserisci la Claude API Key nella sidebar")
        with col_w2:
            st.info("**2️⃣ Carica CSA**\n\nCarica il PDF del Capitolato Speciale d'Appalto")
        with col_w3:
            st.info("**3️⃣ Analizza**\n\nClicca 'Avvia analisi' e ottieni tutti i dati estratti")

        st.markdown("---")
        st.markdown("**Oppure:** usa la **🎮 Modalità Demo** dalla sidebar per esplorare l'app senza PDF.")
        return

    # Alias per compatibilità
    details = csa_data

    # Calcolo importo netto
    importo_base = _parse_importo(csa_data.get("importo_lavori"))
    ribasso_pct = float(st.session_state.get("ribasso_pct", 0.0))
    importo_netto = importo_base * (1 - ribasso_pct / 100) if importo_base > 0 else 0.0

    # Tabs
    tab_dashboard, tab_sintesi, tab_checklist, tab_documenti, tab_calendario, \
    tab_sal_penali, tab_operatori, tab_registri, tab_mappa, tab_pianificazione, \
    tab_log, tab_guida = st.tabs([
        "🏠 Dashboard",
        "📋 Sintesi CSA",
        "✅ Checklist",
        "📄 Documenti",
        "📅 Calendario",
        "💰 SAL e Penali",
        "🏢 Operatori",
        "🗂️ Registri di Cantiere",
        "🗺️ Mappa Fornitori",
        "📋 Pianificazione",
        "📚 Log Attività",
        "❓ Guida",
    ])

    with tab_dashboard:
        _render_dashboard(csa_data, details, importo_netto)

    with tab_sintesi:
        _render_sintesi(csa_data, importo_netto)

    with tab_checklist:
        _render_checklist(csa_data, api_key)

    with tab_documenti:
        _render_documenti(csa_data, api_key)

    with tab_calendario:
        _render_calendario(csa_data, details, importo_netto)

    with tab_sal_penali:
        st.header("💰 SAL e Penali")

        st.subheader("📊 Contabilità SAL")
        _render_contabilita_sal(
            csa_data=csa_data,
            details=details,
            importo_netto=importo_netto,
            results_dir=RESULTS_DIR,
            salva_fn=_salva_stato_cantiere,
        )

        st.divider()

        st.subheader("📉 Penali e Revisione Prezzi")
        _render_penali(csa_data, details, importo_netto)

    with tab_operatori:
        limite_subaffid_10 = importo_netto * 0.10
        st.info(
            f"🔒 **Limite Art. 122 D.Lgs. 36/2023**: Subaffidamenti max **€ {limite_subaffid_10:,.0f}** (10% importo netto)",
            icon="⚠️",
        )

        render_operatori_tab(
            csa_data=csa_data,
            importo_base=importo_netto,
            salva_fn=_salva_stato_cantiere,
            results_dir=RESULTS_DIR,
        )

        # Verifica post-render
        oe_check = st.session_state.get("operatori_economici", {})
        subs_check = oe_check.get("subappaltatori", [])
        tot_subaffid_check = sum(float(s.get("importo", 0) or 0) for s in subs_check if s.get("tipo") == "subaffidamento")
        if tot_subaffid_check > limite_subaffid_10:
            st.error(
                f"❌ **SUPERA LIMITE**: Importo subaffidamenti € {tot_subaffid_check:,.0f}"
                f" > limite € {limite_subaffid_10:,.0f}",
                icon="🚨",
            )

    with tab_registri:
        render_registri_tab(
            csa_data=csa_data,
            details=details,
            results_dir=RESULTS_DIR,
            salva_fn=_salva_stato_cantiere,
            importo_netto=importo_netto,
            include_verbali=False,
            include_contabilita=False,
            include_schede_accettazione=True,
        )

    with tab_mappa:
        _render_mappa(csa_data)

    with tab_pianificazione:
        render_pianificazione_tab(
            csa_data=csa_data,
            api_key=api_key,
            salva_fn=_salva_stato_cantiere,
            results_dir=RESULTS_DIR,
        )

    with tab_log:
        render_log_tab(csa_data=csa_data, salva_fn=_salva_stato_cantiere)

    with tab_guida:
        _render_guida()


if __name__ == "__main__":
    main()
