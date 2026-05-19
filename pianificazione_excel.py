"""
pianificazione_excel.py
═══════════════════════════════════════════════════════════════════════════════
Modulo per la generazione del file Excel di Pianificazione Tempi/Costi.

Incrocia:
  - Fasi del Cronoprogramma (Gantt)
  - Voci del Computo Metrico Estimativo (CME)
  - Oneri per la Sicurezza (non soggetti a ribasso)

Output: file .xlsx con 3 fogli:
  1. Cruscotto per Fase     (dettaglio voci per fase)
  2. Riepilogo per Fase     (tabella sintetica con totali)
  3. WBS Computo Metrico    (struttura gerarchica con %)

Utilizzo in app.py:
  from pianificazione_excel import genera_excel_pianificazione
  buffer = genera_excel_pianificazione(fasi, voci_cme, voci_sicurezza, nome_progetto)
  st.download_button("Scarica Excel", buffer, file_name="Pianificazione.xlsx", ...)

Utilizzo standalone (test):
  python pianificazione_excel.py
  genera Casa_Sgancio_Pianificazione_TEST.xlsx nella cartella corrente
═══════════════════════════════════════════════════════════════════════════════
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── COSTANTI COLORI ─────────────────────────────────────────────────────────
C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_FASE_BG   = "2E75B6"
C_FASE_FG   = "FFFFFF"
C_GROUP_BG  = "BDD7EE"
C_CAT_BG    = "DEEAF1"
C_ITEM_BG   = "FFFFFF"
C_ALT_BG    = "F2F7FB"
C_TOT_BG    = "1F4E79"
C_TOT_FG    = "FFFFFF"
C_SEC_BG    = "C00000"
C_SEC_FG    = "FFFFFF"
C_SEC_ITEM  = "FFE7E7"
C_SEC_ALT   = "FFF2F2"
C_NOTE_BG   = "FFF2CC"
C_NOTE_FG   = "7B0000"

EUR = '#,##0.00 "€"'
PCT = '0.00%'


# ─── HELPERS ─────────────────────────────────────────────────────────────────
def _fill(h):
    return PatternFill("solid", start_color=h, fgColor=h)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")

_thin  = Side(style="thin",   color="B0B0B0")
_thick = Side(style="medium", color="1F4E79")

def _b():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _bt():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thick)


# ─── API PUBBLICA ─────────────────────────────────────────────────────────────
def genera_excel_pianificazione(
    fasi: list,
    voci_cme: list,
    voci_sicurezza: list,
    nome_progetto: str = "Progetto"
) -> io.BytesIO:
    """
    Genera il file Excel di pianificazione.

    Struttura dati attesa:

    fasi = [
        {"nome": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "durata_giorni": 5,
         "settimana_inizio": 1,
         "settimana_fine": 1},
        ...
    ]

    voci_cme = [
        {"numero": "1/1",
         "categoria": "DEMOLIZIONI",
         "descrizione": "Rimozione manto di copertura",
         "quantita": 180.0,
         "unita_misura": "mq",
         "prezzo_unitario": 19.66,
         "importo_totale": 3538.80,
         "fase_associata": "FASE 2 – OPERE DI DEMOLIZIONE"},
        ...
    ]

    voci_sicurezza = [
        {"numero": "1",
         "descrizione": "Ponteggio telaio prefabbricato h>10m",
         "quantita": 648.0,
         "unita_misura": "mq",
         "prezzo_unitario": 21.53,
         "importo_totale": 13951.44,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        ...
    ]

    Returns:
        io.BytesIO pronto per st.download_button()
    """
    wb = Workbook()

    tot_cme     = sum(v.get("importo_totale", 0) or 0 for v in voci_cme)
    tot_sic     = sum(v.get("importo_totale", 0) or 0 for v in voci_sicurezza)
    tot_appalto = tot_cme + tot_sic

    _sheet_cruscotto(wb, fasi, voci_cme, voci_sicurezza,
                     nome_progetto, tot_cme, tot_sic, tot_appalto)
    _sheet_riepilogo(wb, fasi, voci_cme, voci_sicurezza,
                     nome_progetto, tot_cme, tot_sic, tot_appalto)
    _sheet_wbs(wb, voci_cme, voci_sicurezza, tot_cme, tot_sic, tot_appalto)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── FOGLIO 1: CRUSCOTTO PER FASE ────────────────────────────────────────────
def _sheet_cruscotto(wb, fasi, voci_cme, voci_sicurezza,
                     nome_progetto, tot_cme, tot_sic, tot_appalto):
    ws = wb.active
    ws.title = "Cruscotto per Fase"

    for col, w in zip("ABCDEFGHIJ",
                      [6, 50, 10, 8, 7, 7, 9, 13, 14, 16]):
        ws.column_dimensions[col].width = w

    # titolo progetto
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1,
                value=f"PIANIFICAZIONE TEMPI E COSTI — {nome_progetto.upper()}")
    c.font = _font(bold=True, color=C_HEADER_FG, size=12)
    c.fill = _fill(C_HEADER_BG)
    c.alignment = _center()
    c.border = _b()

    # intestazione colonne
    hdrs = ["N°", "DESIGNAZIONE DEI LAVORI", "Q.TÀ", "U.M.",
            "S.INI", "S.FIN", "GG", "P.U. (€)",
            "IMP. CME (€)", "IMP. SIC. (€)"]
    ws.row_dimensions[2].height = 36
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, color=C_HEADER_FG, size=10)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border = _b()

    row = 3
    for fase in fasi:
        nome  = fase.get("nome", "")
        dur   = fase.get("durata_giorni", 0)
        sw    = fase.get("settimana_inizio", 1)
        ew    = fase.get("settimana_fine", 1)

        # titolo fase
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=8)
        c = ws.cell(row=row, column=1, value=nome)
        c.font = _font(bold=True, color=C_FASE_FG, size=11)
        c.fill = _fill(C_FASE_BG)
        c.alignment = _left()
        c.border = _b()
        for col, val in [(9, f"Sett.{sw}÷{ew}"), (10, f"{dur} gg")]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(bold=True, color=C_FASE_FG)
            cell.fill = _fill(C_FASE_BG)
            cell.alignment = _center()
            cell.border = _b()
        row += 1

        # CME
        fase_cme = [v for v in voci_cme
                    if v.get("fase_associata", "") == nome]
        if fase_cme:
            ws.row_dimensions[row].height = 16
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row,   end_column=10)
            h = ws.cell(row=row, column=1,
                        value="   ▸  LAVORI A MISURA – COMPUTO METRICO ESTIMATIVO")
            h.font = _font(bold=True, color="1F4E79", size=9)
            h.fill = _fill(C_GROUP_BG)
            h.alignment = _left()
            h.border = _b()
            row += 1

            cats = list(dict.fromkeys(
                v.get("categoria", "Generale") for v in fase_cme))
            for cat in cats:
                cat_items = [v for v in fase_cme
                             if v.get("categoria", "Generale") == cat]
                ws.row_dimensions[row].height = 15
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row,   end_column=10)
                ch = ws.cell(row=row, column=1,
                             value=f"      › {cat.upper()}")
                ch.font = _font(bold=True, color="1F4E79", size=9)
                ch.fill = _fill(C_CAT_BG)
                ch.alignment = _left()
                ch.border = _b()
                row += 1

                for idx, v in enumerate(cat_items):
                    ws.row_dimensions[row].height = 14
                    bg = C_ALT_BG if idx % 2 == 1 else C_ITEM_BG
                    qty = v.get("quantita", "")
                    pu  = v.get("prezzo_unitario", "")
                    tot = v.get("importo_totale", "")
                    vals = [v.get("numero",""), v.get("descrizione",""),
                            qty, v.get("unita_misura",""),
                            sw, ew, dur,
                            pu if pu else "",
                            tot if tot else "", ""]
                    for ci, val in enumerate(vals, 1):
                        cell = ws.cell(row=row, column=ci, value=val)
                        cell.fill = _fill(bg)
                        cell.border = _b()
                        cell.font = _font(size=9)
                        if ci == 1:   cell.alignment = _center()
                        elif ci == 2: cell.alignment = _left()
                        elif ci in (3,4,5,6,7): cell.alignment = _center()
                        elif ci in (8,9):
                            cell.alignment = _right()
                            if isinstance(val, (int, float)) and val:
                                cell.number_format = EUR
                    row += 1

        # Sicurezza
        fase_sic = [v for v in voci_sicurezza
                    if v.get("fase_associata", "") == nome]
        if fase_sic:
            ws.row_dimensions[row].height = 16
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row,   end_column=10)
            sh = ws.cell(row=row, column=1,
                         value="   🔴  ONERI PER LA SICUREZZA – NON SOGGETTI A RIBASSO")
            sh.font = _font(bold=True, color=C_SEC_FG, size=9)
            sh.fill = _fill(C_SEC_BG)
            sh.alignment = _left()
            sh.border = _b()
            row += 1

            for idx, v in enumerate(fase_sic):
                ws.row_dimensions[row].height = 14
                bg = C_SEC_ALT if idx % 2 == 1 else C_SEC_ITEM
                qty = v.get("quantita", "")
                pu  = v.get("prezzo_unitario", "")
                tot = v.get("importo_totale", "")
                vals = [v.get("numero",""), v.get("descrizione",""),
                        qty, v.get("unita_misura",""),
                        sw, ew, dur,
                        pu if pu else "", "",
                        tot if tot else ""]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.fill = _fill(bg)
                    cell.border = _b()
                    cell.font = _font(size=9, color="7B0000")
                    if ci == 1:   cell.alignment = _center()
                    elif ci == 2: cell.alignment = _left()
                    elif ci in (3,4,5,6,7): cell.alignment = _center()
                    elif ci in (8,9,10):
                        cell.alignment = _right()
                        if isinstance(val, (int, float)) and val:
                            cell.number_format = EUR
                row += 1

        # totale fase
        ws.row_dimensions[row].height = 18
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=8)
        lbl = ws.cell(row=row, column=1, value=f"TOTALE  {nome}")
        lbl.font = _font(bold=True, color=C_TOT_FG, size=10)
        lbl.fill = _fill(C_TOT_BG)
        lbl.alignment = _right()
        lbl.border = _bt()

        cme_f = sum(v.get("importo_totale",0) or 0
                    for v in voci_cme if v.get("fase_associata","") == nome)
        sic_f = sum(v.get("importo_totale",0) or 0
                    for v in voci_sicurezza if v.get("fase_associata","") == nome)

        for ci, val, color in [(9, cme_f, C_TOT_FG), (10, sic_f, "FFCCCC")]:
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _font(bold=True, color=color, size=10)
            cell.fill = _fill(C_TOT_BG)
            cell.alignment = _right()
            cell.number_format = EUR
            cell.border = _bt()
        row += 2

    # grand total CME
    ws.row_dimensions[row].height = 24
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=8)
    ws.cell(row=row, column=1,
            value="TOTALE GENERALE LAVORI A MISURA").font = _font(
                bold=True, color=C_HEADER_FG, size=12)
    ws.cell(row=row, column=1).fill = _fill(C_HEADER_BG)
    ws.cell(row=row, column=1).alignment = _right()
    ws.cell(row=row, column=1).border = _bt()
    for ci, val, color in [(9, tot_cme, C_HEADER_FG),
                           (10, tot_sic, "FFCCCC")]:
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = _font(bold=True, color=color, size=12)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _right()
        cell.number_format = EUR
        cell.border = _bt()
    row += 1

    # totale complessivo
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=8)
    ws.cell(row=row, column=1,
            value="TOTALE COMPLESSIVO (CME + Oneri Sicurezza)").font = _font(
                bold=True, color=C_SEC_FG, size=11)
    ws.cell(row=row, column=1).fill = _fill(C_SEC_BG)
    ws.cell(row=row, column=1).alignment = _right()
    ws.cell(row=row, column=1).border = _bt()
    cell = ws.cell(row=row, column=9, value=tot_appalto)
    cell.font = _font(bold=True, color=C_SEC_FG, size=11)
    cell.fill = _fill(C_SEC_BG)
    cell.alignment = _right()
    cell.number_format = EUR
    cell.border = _bt()
    ws.cell(row=row, column=10).fill = _fill(C_SEC_BG)
    ws.cell(row=row, column=10).border = _bt()


# ─── FOGLIO 2: RIEPILOGO PER FASE ────────────────────────────────────────────
def _sheet_riepilogo(wb, fasi, voci_cme, voci_sicurezza,
                     nome_progetto, tot_cme, tot_sic, tot_appalto):
    ws = wb.create_sheet("Riepilogo per Fase")
    for col, w in zip("ABCDEF", [42, 12, 14, 16, 16, 10]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1,
                value=f"RIEPILOGO PER FASE — {nome_progetto.upper()}")
    t.font = _font(bold=True, color=C_HEADER_FG, size=12)
    t.fill = _fill(C_HEADER_BG)
    t.alignment = _center()
    t.border = _b()

    hdrs = ["FASE", "DURATA (gg)", "SETTIMANE",
            "IMPORTO CME (€)", "ONERI SIC. (€)", "% CME"]
    ws.row_dimensions[2].height = 28
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border = _b()

    for fi, fase in enumerate(fasi):
        nome = fase.get("nome", "")
        dur  = fase.get("durata_giorni", 0)
        sw   = fase.get("settimana_inizio", 1)
        ew   = fase.get("settimana_fine", 1)
        cme_f = sum(v.get("importo_totale",0) or 0
                    for v in voci_cme if v.get("fase_associata","") == nome)
        sic_f = sum(v.get("importo_totale",0) or 0
                    for v in voci_sicurezza if v.get("fase_associata","") == nome)
        pct   = cme_f / tot_cme if tot_cme else 0

        r  = fi + 3
        bg = C_ALT_BG if fi % 2 == 0 else C_ITEM_BG
        ws.row_dimensions[r].height = 20
        for ci, (val, fmt, aln) in enumerate(zip(
            [nome, dur, f"S{sw}÷S{ew}",
             cme_f, sic_f if sic_f else "—", pct],
            [None, None, None, EUR, EUR if sic_f else None, PCT],
            [_left, _center, _center, _right, _right, _center]
        ), 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = _font(bold=(ci == 1))
            cell.fill = _fill(bg)
            cell.border = _b()
            cell.alignment = aln()
            if fmt: cell.number_format = fmt

    tr = len(fasi) + 3
    ws.row_dimensions[tr].height = 22
    dur_tot = sum(f.get("durata_giorni", 0) for f in fasi)
    sw_min  = min(f.get("settimana_inizio", 1) for f in fasi) if fasi else 1
    ew_max  = max(f.get("settimana_fine",   1) for f in fasi) if fasi else 1
    for ci, (val, fmt, aln) in enumerate(zip(
        ["TOTALE COMPLESSIVO", dur_tot, f"S{sw_min}÷S{ew_max}",
         tot_cme, tot_sic, 1.0],
        [None, None, None, EUR, EUR, PCT],
        [_left, _center, _center, _right, _right, _center]
    ), 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.border = _bt()
        c.alignment = aln()
        if fmt: c.number_format = fmt

    # note
    nr = tr + 1
    ws.row_dimensions[nr].height = 16
    ws.merge_cells(f"A{nr}:F{nr}")
    n = ws.cell(row=nr, column=1,
                value=f"⚠️  Gli Oneri per la Sicurezza (€ {tot_sic:,.2f}) "
                      f"NON sono soggetti a ribasso d'asta (art. 41 D.Lgs. 36/2023).")
    n.font = _font(bold=True, color=C_NOTE_FG, size=9)
    n.fill = _fill(C_NOTE_BG)
    n.alignment = _left()
    n.border = _b()

    nr2 = nr + 1
    ws.row_dimensions[nr2].height = 14
    ws.merge_cells(f"A{nr2}:F{nr2}")
    n2 = ws.cell(row=nr2, column=1,
                 value=f"   Base d'asta soggetta a ribasso: € {tot_cme:,.2f}  |  "
                       f"Sicurezza non ribassabile: € {tot_sic:,.2f}  |  "
                       f"TOTALE APPALTO: € {tot_appalto:,.2f}")
    n2.font = _font(size=9)
    n2.fill = _fill(C_NOTE_BG)
    n2.alignment = _left()
    n2.border = _b()


# ─── FOGLIO 3: WBS COMPUTO METRICO ───────────────────────────────────────────
def _sheet_wbs(wb, voci_cme, voci_sicurezza, tot_cme, tot_sic, tot_appalto):
    ws = wb.create_sheet("WBS Computo Metrico")
    for col, w in zip("ABCD", [55, 16, 10, 20]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(["DESCRIZIONE", "IMPORTO (€)", "% TOT.", "NOTE"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border = _b()

    row = 2

    # totale CME
    ws.row_dimensions[row].height = 20
    for ci, (val, fmt, aln) in enumerate(zip(
        [f"TOTALE CME — {len(voci_cme)} VOCI",
         tot_cme,
         tot_cme / tot_appalto if tot_appalto else 0,
         "Soggetto a ribasso"],
        [None, EUR, PCT, None],
        [_left, _right, _center, _center]
    ), 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _font(bold=True, color=C_HEADER_FG)
        c.fill = _fill(C_HEADER_BG)
        c.border = _bt()
        c.alignment = aln()
        if fmt: c.number_format = fmt
    row += 1

    # categorie CME
    cats = list(dict.fromkeys(
        v.get("categoria", "Generale") for v in voci_cme))
    for idx_c, cat in enumerate(cats):
        items  = [v for v in voci_cme
                  if v.get("categoria", "Generale") == cat]
        c_tot  = sum(v.get("importo_totale", 0) or 0 for v in items)
        pct    = c_tot / tot_cme if tot_cme else 0
        bg     = C_ALT_BG if idx_c % 2 == 0 else C_ITEM_BG

        ws.row_dimensions[row].height = 17
        for ci, (val, fmt, aln) in enumerate(zip(
            [f"  › {cat}", c_tot, pct, f"{len(items)} voci"],
            [None, EUR, PCT, None],
            [_left, _right, _center, _center]
        ), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = _font(bold=(ci == 1))
            c.fill = _fill(bg)
            c.border = _b()
            c.alignment = aln()
            if fmt: c.number_format = fmt
        row += 1

    row += 1

    # header sicurezza
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=4)
    sh = ws.cell(row=row, column=1,
                 value="🔴  ONERI PER LA SICUREZZA — NON SOGGETTI A RIBASSO")
    sh.font = _font(bold=True, color=C_SEC_FG, size=10)
    sh.fill = _fill(C_SEC_BG)
    sh.alignment = _left()
    sh.border = _b()
    row += 1

    for idx, v in enumerate(voci_sicurezza):
        ws.row_dimensions[row].height = 15
        bg  = C_SEC_ALT if idx % 2 == 1 else C_SEC_ITEM
        tot = v.get("importo_totale", 0) or 0
        pct = tot / tot_sic if tot_sic else 0
        for ci, (val, fmt, aln) in enumerate(zip(
            [f"  {v.get('numero','')} – {v.get('descrizione','')}",
             tot, pct, "Non ribassabile"],
            [None, EUR, PCT, None],
            [_left, _right, _center, _center]
        ), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = _font(size=9, color="7B0000")
            c.fill = _fill(bg)
            c.border = _b()
            c.alignment = aln()
            if fmt: c.number_format = fmt
        row += 1

    # totale sicurezza
    ws.row_dimensions[row].height = 18
    for ci, (val, fmt, aln) in enumerate(zip(
        ["TOTALE ONERI SICUREZZA",
         tot_sic,
         tot_sic / tot_appalto if tot_appalto else 0,
         "NON RIBASSABILE"],
        [None, EUR, PCT, None],
        [_left, _right, _center, _center]
    ), 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _font(bold=True, color=C_SEC_FG)
        c.fill = _fill(C_SEC_BG)
        c.border = _bt()
        c.alignment = aln()
        if fmt: c.number_format = fmt
    row += 1

    # totale appalto
    ws.row_dimensions[row].height = 22
    for ci, (val, fmt, aln) in enumerate(zip(
        ["TOTALE APPALTO (CME + Oneri Sicurezza)",
         tot_appalto, 1.0, ""],
        [None, EUR, PCT, None],
        [_left, _right, _center, _center]
    ), 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = _font(bold=True, color=C_HEADER_FG, size=11)
        c.fill = _fill(C_HEADER_BG)
        c.border = _bt()
        c.alignment = aln()
        if fmt: c.number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
# TEST STANDALONE
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    fasi_test = [
        {"nome": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "durata_giorni": 5, "settimana_inizio": 1, "settimana_fine": 1},
        {"nome": "FASE 2 – OPERE DI DEMOLIZIONE",
         "durata_giorni": 8, "settimana_inizio": 2, "settimana_fine": 3},
        {"nome": "FASE 3 – OPERE ESTERNE",
         "durata_giorni": 44, "settimana_inizio": 4, "settimana_fine": 14},
        {"nome": "FASE 4 – OPERE INTERNE",
         "durata_giorni": 144, "settimana_inizio": 15, "settimana_fine": 50},
    ]

    voci_cme_test = [
        {"numero": "1/1", "categoria": "Demolizioni e rimozioni",
         "descrizione": "Rimozione manto di copertura",
         "quantita": 180.0, "unita_misura": "mq",
         "prezzo_unitario": 19.66, "importo_totale": 3538.80,
         "fase_associata": "FASE 2 – OPERE DI DEMOLIZIONE"},
        {"numero": "20/84", "categoria": "Demolizioni e rimozioni",
         "descrizione": "Rimozione-bonifica coperture in cemento-amianto",
         "quantita": 1.0, "unita_misura": "corpo",
         "prezzo_unitario": 1528.29, "importo_totale": 1528.29,
         "fase_associata": "FASE 2 – OPERE DI DEMOLIZIONE"},
        {"numero": "21/69", "categoria": "Opere strutturali",
         "descrizione": "Orditura tetto lamellare schema complesso",
         "quantita": 2.25, "unita_misura": "mc",
         "prezzo_unitario": 2894.06, "importo_totale": 6511.64,
         "fase_associata": "FASE 3 – OPERE ESTERNE"},
        {"numero": "32/80", "categoria": "Impermeabilizzazioni e lattoneria",
         "descrizione": "Pannello sandwich copertura sp.80mm",
         "quantita": 230.0, "unita_misura": "mq",
         "prezzo_unitario": 50.01, "importo_totale": 11502.30,
         "fase_associata": "FASE 3 – OPERE ESTERNE"},
        {"numero": "80/42", "categoria": "Infissi esterni – sostituzione",
         "descrizione": "Serramento PVC battente 70mm",
         "quantita": 28.49, "unita_misura": "mq",
         "prezzo_unitario": 410.28, "importo_totale": 11688.88,
         "fase_associata": "FASE 4 – OPERE INTERNE"},
        {"numero": "91/91", "categoria": "Impianti tecnologici",
         "descrizione": "Climatizzatore DC inverter mono-split 10,5kW",
         "quantita": 1.0, "unita_misura": "cad",
         "prezzo_unitario": 7030.97, "importo_totale": 7030.97,
         "fase_associata": "FASE 4 – OPERE INTERNE"},
    ]

    voci_sic_test = [
        {"numero": "1", "descrizione": "Ponteggio telaio prefabbricato h>10m (1° mese)",
         "quantita": 648.0, "unita_misura": "mq",
         "prezzo_unitario": 21.53, "importo_totale": 13951.44,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        {"numero": "2", "descrizione": "Ponteggio – maggiorazione mesi successivi",
         "quantita": 1296.0, "unita_misura": "m²/mese",
         "prezzo_unitario": 3.07, "importo_totale": 3978.72,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        {"numero": "3", "descrizione": "Schermatura antipolvere teli PE 240g/mq",
         "quantita": 648.0, "unita_misura": "mq",
         "prezzo_unitario": 1.68, "importo_totale": 1088.64,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        {"numero": "4", "descrizione": "Ponte su cavalletti h≤4m (controsoffitto)",
         "quantita": 100.0, "unita_misura": "mq",
         "prezzo_unitario": 2.71, "importo_totale": 271.00,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        {"numero": "5", "descrizione": "Bagno chimico portatile (costo mensile)",
         "quantita": 1.0, "unita_misura": "cad",
         "prezzo_unitario": 199.87, "importo_totale": 199.87,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
        {"numero": "6-8", "descrizione": "Cartelli divieto/obbligo/pericolo",
         "quantita": 6.0, "unita_misura": "cad",
         "prezzo_unitario": 5.72, "importo_totale": 34.32,
         "fase_associata": "FASE 1 – ORGANIZZAZIONE CANTIERE",
         "non_ribassabile": True},
    ]

    buf = genera_excel_pianificazione(
        fasi_test, voci_cme_test, voci_sic_test,
        "Casa di Sgancio – Pescara (PE)"
    )
    out = "Casa_Sgancio_Pianificazione_TEST.xlsx"
    with open(out, "wb") as f:
        f.write(buf.read())

    tot_c = sum(v["importo_totale"] for v in voci_cme_test)
    tot_s = sum(v["importo_totale"] for v in voci_sic_test)
    print(f"✅  File generato: {out}")
    print(f"   CME:       € {tot_c:,.2f}")
    print(f"   Sicurezza: € {tot_s:,.2f}")
    print(f"   Totale:    € {tot_c + tot_s:,.2f}")
